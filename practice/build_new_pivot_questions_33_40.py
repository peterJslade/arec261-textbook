#!/usr/bin/env python3
"""Build the answer workbooks for the defined Module 1 Questions 33–40.

Question 37 is intentionally omitted because it is not present in
``module01_bank.qmd``.
"""

import csv
import os
import shutil
import sys
import zipfile
from collections import defaultdict
from pathlib import Path

sys.path.insert(0, os.path.expanduser("~/.claude/skills/excel-pivot-tables/scripts"))

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
import xlsx_pivot


ROOT = Path(__file__).resolve().parents[1]
RM_SOURCE = ROOT / "practice/data/rm_yields_1990_2025.csv"
MB_SOURCE = ROOT / "practice/data/mb_wheat_reported_2020_2025.csv"
ANSWER_DIR = ROOT / "practice/answers"
OUTPUT_DIR = ROOT / "outputs/module01-new-pivot-answers"

GREEN = "4A7C59"
PALE_GREEN = "E2F0D9"
YELLOW = "FFF2CC"
TEXT = "24302A"
THIN_GREEN = Side(style="thin", color="A9C4AE")

RM_HEADERS = ["Year", "RM", "Crop", "Yield", "Unit"]
MB_HEADERS = [
    "Year",
    "Municipality",
    "Variety",
    "Farms",
    "Acres",
    "Yield_bu_ac",
    "Reported",
]


def read_rows(path, numeric_fields):
    with path.open(newline="", encoding="utf-8-sig") as handle:
        rows = list(csv.DictReader(handle))
    for row in rows:
        for field in numeric_fields:
            if row[field] != "":
                value = float(row[field])
                row[field] = int(value) if value.is_integer() else value
    return rows


def write_source_workbook(path, rows, headers, helper_rows, helper_start_col):
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    ws.append(headers)
    for row in rows:
        ws.append([row[h] for h in headers])

    for offset, header in enumerate(headers):
        ws.cell(1, helper_start_col + offset, header)
    for row_number, row in enumerate(helper_rows, 2):
        for offset, header in enumerate(headers):
            ws.cell(row_number, helper_start_col + offset, row[header])

    style_data_sheet(ws, headers, len(rows) + 1)
    for col in range(helper_start_col, helper_start_col + len(headers)):
        ws.cell(1, col).fill = PatternFill("solid", fgColor=GREEN)
        ws.cell(1, col).font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
        ws.column_dimensions[ws.cell(1, col).column_letter].hidden = True
    wb.save(path)


def style_data_sheet(ws, headers, last_row):
    for cell in ws[1][: len(headers)]:
        cell.fill = PatternFill("solid", fgColor=GREEN)
        cell.font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{ws.cell(1, len(headers)).column_letter}{last_row}"
    widths = {
        "Year": 11,
        "RM": 10,
        "Crop": 19,
        "Yield": 12,
        "Unit": 11,
        "Municipality": 24,
        "Variety": 42,
        "Farms": 11,
        "Acres": 14,
        "Yield_bu_ac": 15,
        "Reported": 12,
    }
    for index, header in enumerate(headers, 1):
        ws.column_dimensions[ws.cell(1, index).column_letter].width = widths[header]
    ws.sheet_properties.tabColor = GREEN


def format_pivot_sheet(ws, title):
    ws["A1"] = title
    ws["A1"].font = Font(name="Arial", size=14, bold=True, color=GREEN)
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = GREEN
    ws.freeze_panes = "A4"


def highlight_row(ws, row, first_col, last_col):
    for col in range(first_col, last_col + 1):
        ws.cell(row, col).fill = PatternFill("solid", fgColor=YELLOW)


def add_answer_box(ws, label_range, answer_range, answer):
    ws.merge_cells(label_range)
    label = ws[label_range.split(":")[0]]
    label.value = "Answer to part (b)"
    label.fill = PatternFill("solid", fgColor=GREEN)
    label.font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
    label.alignment = Alignment(horizontal="center")

    ws.merge_cells(answer_range)
    cell = ws[answer_range.split(":")[0]]
    cell.value = answer
    cell.fill = PatternFill("solid", fgColor=YELLOW)
    cell.font = Font(name="Arial", size=11, bold=True, color=TEXT)
    cell.alignment = Alignment(vertical="center", wrap_text=True)
    cell.border = Border(
        left=THIN_GREEN, right=THIN_GREEN, top=THIN_GREEN, bottom=THIN_GREEN
    )


def disable_refresh_on_open(path):
    temp = path.with_suffix(".patched.xlsx")
    with zipfile.ZipFile(path, "r") as source, zipfile.ZipFile(
        temp, "w", zipfile.ZIP_DEFLATED
    ) as target:
        for item in source.infolist():
            data = source.read(item.filename)
            if item.filename.startswith("xl/pivotCache/pivotCacheDefinition"):
                data = data.replace(b'refreshOnLoad="1"', b'refreshOnLoad="0"')
            target.writestr(item, data)
    temp.replace(path)


def aggregate(rows, group_field, value_field, method):
    grouped = defaultdict(list)
    for row in rows:
        grouped[row[group_field]].append(float(row[value_field]))
    if method == "average":
        return {key: sum(values) / len(values) for key, values in grouped.items()}
    if method == "sum":
        return {key: sum(values) for key, values in grouped.items()}
    if method == "count":
        return {key: len(values) for key, values in grouped.items()}
    raise ValueError(method)


def add_pivot(path, helper_start_col, helper_rows, headers, **kwargs):
    start_letter = chr(64 + helper_start_col)
    end_letter = chr(64 + helper_start_col + len(headers) - 1)
    xlsx_pivot.add_pivot_table(
        path,
        source_sheet="Data",
        source_ref=f"{start_letter}1:{end_letter}{len(helper_rows) + 1}",
        dest_sheet="PivotTable",
        dest_cell="A3",
        output_path=path,
        **kwargs,
    )


def with_sort_order(order):
    original = xlsx_pivot._sort_key
    rank = {value: index for index, value in enumerate(order)}
    xlsx_pivot._sort_key = (
        lambda value: (rank[value],) if value in rank else original(value)
    )
    return original


def finish(path, title):
    wb = load_workbook(path)
    ws = wb["PivotTable"]
    format_pivot_sheet(ws, title)
    return wb, ws


def save_finished(wb, path):
    wb.active = wb.sheetnames.index("PivotTable")
    wb.save(path)
    disable_refresh_on_open(path)
    shutil.copy2(path, OUTPUT_DIR / path.name)
    print(path)


def find_row(ws, value, column=1):
    return next(
        row
        for row in range(1, ws.max_row + 1)
        if str(ws.cell(row, column).value) == str(value)
    )


def find_col(ws, value, max_rows=6):
    for row in range(1, min(ws.max_row, max_rows) + 1):
        for col in range(1, ws.max_column + 1):
            if str(ws.cell(row, col).value) == str(value):
                return col
    raise RuntimeError(f"Could not find PivotTable column {value!r}")


def build_q33(rm_rows):
    path = ANSWER_DIR / "q033.xlsx"
    helper = [row for row in rm_rows if row["Year"] > 2015]
    crops = sorted({row["Crop"] for row in helper})
    write_source_workbook(path, rm_rows, RM_HEADERS, helper, 7)
    original = with_sort_order(crops + list(range(2016, 2026)))
    add_pivot(
        path,
        7,
        helper,
        RM_HEADERS,
        rows=["Crop"],
        cols=["Year"],
        values=[{
            "field": "Yield",
            "agg": "average",
            "name": "Average Yield",
            "num_format": "0.0",
        }],
        name="Question33Pivot",
    )
    xlsx_pivot._sort_key = original
    wb, ws = finish(path, "Question 33 — Average Crop Yield by Year, 2016–2025")
    col_2021 = find_col(ws, 2021)
    header_row = find_row(ws, "Crop")
    grand_row = find_row(ws, "Grand Total")
    for row in range(header_row + 1, grand_row):
        ws.cell(row, col_2021).fill = PatternFill("solid", fgColor=YELLOW)
    answer_row = grand_row + 3
    add_answer_box(
        ws,
        f"A{answer_row}:L{answer_row}",
        f"A{answer_row + 1}:L{answer_row + 2}",
        "All eight crops—Barley, Canola, Durum, Flax, Lentils, Oats, Peas, and Spring Wheat—had their lowest average yield in 2021.",
    )
    ws.row_dimensions[answer_row + 1].height = 27
    ws.row_dimensions[answer_row + 2].height = 27
    save_finished(wb, path)


def build_q34(rm_rows):
    path = ANSWER_DIR / "q034.xlsx"
    helper = [row for row in rm_rows if row["Year"] == 2023]
    counts = aggregate(helper, "Crop", "RM", "count")
    order = sorted(counts, key=lambda crop: (counts[crop], crop))
    write_source_workbook(path, rm_rows, RM_HEADERS, helper, 7)
    original = with_sort_order(order)
    add_pivot(
        path,
        7,
        helper,
        RM_HEADERS,
        rows=["Crop"],
        values=[{
            "field": "RM",
            "agg": "count",
            "name": "RMs Reporting",
            "num_format": "#,##0",
        }],
        name="Question34Pivot",
    )
    xlsx_pivot._sort_key = original
    wb, ws = finish(path, "Question 34 — Number of RMs Reporting Each Crop in 2023")
    header_row = find_row(ws, "Crop")
    highlight_row(ws, header_row + 1, 1, 2)
    add_answer_box(ws, "D3:H3", "D4:H5", "Flax was reported in the fewest RMs: 164.")
    ws.row_dimensions[4].height = 26
    ws.row_dimensions[5].height = 26
    save_finished(wb, path)


def build_q35(rm_rows):
    path = ANSWER_DIR / "q035.xlsx"
    helper = [
        row
        for row in rm_rows
        if row["Year"] >= 2000 and row["Crop"] in {"Durum", "Spring Wheat"}
    ]
    write_source_workbook(path, rm_rows, RM_HEADERS, helper, 7)
    original = with_sort_order(list(range(2000, 2026)) + ["Durum", "Spring Wheat"])
    add_pivot(
        path,
        7,
        helper,
        RM_HEADERS,
        rows=["Year"],
        cols=["Crop"],
        values=[{
            "field": "Yield",
            "agg": "average",
            "name": "Average Yield (bu/ac)",
            "num_format": "0.0",
        }],
        name="Question35Pivot",
    )
    xlsx_pivot._sort_key = original
    wb, ws = finish(path, "Question 35 — Average Durum and Spring Wheat Yields Since 2000")
    header_row = find_row(ws, "Year")
    for year in list(range(2000, 2005)) + list(range(2021, 2026)):
        row = find_row(ws, year)
        highlight_row(ws, row, 1, ws.max_column)
    add_answer_box(
        ws,
        "F3:J3",
        "F4:J7",
        "Durum averaged 26.1 bu/ac in 2000–2004 and 36.1 in 2021–2025. Spring Wheat rose from 26.0 to 43.9 bu/ac. Both crops show an upward trend, especially Spring Wheat.",
    )
    for col in ("F", "G", "H", "I", "J"):
        ws.column_dimensions[col].width = 16
    for row in range(4, 8):
        ws.row_dimensions[row].height = 25
    save_finished(wb, path)


def build_q36(mb_rows):
    path = ANSWER_DIR / "q036.xlsx"
    helper = [row for row in mb_rows if row["Year"] == 2024]
    averages = aggregate(helper, "Variety", "Yield_bu_ac", "average")
    order = sorted(averages, key=lambda variety: (-averages[variety], variety))
    write_source_workbook(path, mb_rows, MB_HEADERS, helper, 9)
    original = with_sort_order(order)
    add_pivot(
        path,
        9,
        helper,
        MB_HEADERS,
        rows=["Variety"],
        values=[
            {
                "field": "Yield_bu_ac",
                "agg": "average",
                "name": "Average Yield (bu/ac)",
                "num_format": "0.0",
            },
            {
                "field": "Municipality",
                "agg": "count",
                "name": "Municipalities Reporting",
                "num_format": "#,##0",
            },
        ],
        name="Question36Pivot",
    )
    xlsx_pivot._sort_key = original
    wb, ws = finish(path, "Question 36 — 2024 Yield and Reporting Municipalities by Variety")
    header_row = find_row(ws, "Variety")
    highlight_row(ws, header_row + 1, 1, 3)
    add_answer_box(
        ws,
        "E3:I3",
        "E4:I6",
        "SY GABBRO <SYNGENTA> had the highest average yield, 83.4 bu/ac, based on 1 municipality.",
    )
    for col in ("E", "F", "G", "H", "I"):
        ws.column_dimensions[col].width = 16
    for row in range(4, 7):
        ws.row_dimensions[row].height = 25
    save_finished(wb, path)


def build_q37(mb_rows):
    path = ANSWER_DIR / "q037.xlsx"
    helper = list(mb_rows)
    totals = aggregate(helper, "Variety", "Acres", "sum")
    order = sorted(totals, key=lambda variety: (-totals[variety], variety))
    write_source_workbook(path, mb_rows, MB_HEADERS, helper, 9)
    original = with_sort_order(order)
    add_pivot(
        path,
        9,
        helper,
        MB_HEADERS,
        rows=["Variety"],
        values=[{
            "field": "Acres",
            "agg": "sum",
            "name": "Total Acres",
            "num_format": "#,##0",
        }],
        name="Question37Pivot",
    )
    xlsx_pivot._sort_key = original
    wb, ws = finish(path, "Question 37 — Total Acres by Variety, 2020–2025")
    header_row = find_row(ws, "Variety")
    highlight_row(ws, header_row + 1, 1, 2)
    add_answer_box(
        ws,
        "D3:H3",
        "D4:H6",
        "AAC BRANDON (BW 932) was planted on the most acres: 6,741,112 acres over 2020–2025.",
    )
    for col in ("D", "E", "F", "G", "H"):
        ws.column_dimensions[col].width = 16
    for row in range(4, 7):
        ws.row_dimensions[row].height = 25
    save_finished(wb, path)


def build_q38(mb_rows):
    path = ANSWER_DIR / "q038.xlsx"
    helper = list(mb_rows)
    write_source_workbook(path, mb_rows, MB_HEADERS, helper, 9)
    original = with_sort_order(list(range(2020, 2026)))
    add_pivot(
        path,
        9,
        helper,
        MB_HEADERS,
        rows=["Year"],
        values=[{
            "field": "Yield_bu_ac",
            "agg": "average",
            "name": "Average Yield (bu/ac)",
            "num_format": "0.0",
        }],
        name="Question38Pivot",
    )
    xlsx_pivot._sort_key = original
    wb, ws = finish(path, "Question 38 — Manitoba Average Wheat Yield by Year")
    highlight_row(ws, find_row(ws, 2021), 1, 2)
    add_answer_box(ws, "D3:H3", "D4:H5", "2021 was the worst year, averaging 49.6 bu/ac.")
    ws.row_dimensions[4].height = 26
    ws.row_dimensions[5].height = 26
    save_finished(wb, path)


def build_q39(mb_rows):
    path = ANSWER_DIR / "q039.xlsx"
    helper = [row for row in mb_rows if row["Year"] == 2024]
    averages = aggregate(helper, "Municipality", "Yield_bu_ac", "average")
    order = sorted(averages, key=lambda municipality: (-averages[municipality], municipality))
    write_source_workbook(path, mb_rows, MB_HEADERS, helper, 9)
    original = with_sort_order(order)
    add_pivot(
        path,
        9,
        helper,
        MB_HEADERS,
        rows=["Municipality"],
        values=[{
            "field": "Yield_bu_ac",
            "agg": "average",
            "name": "Average Yield (bu/ac)",
            "num_format": "0.0",
        }],
        name="Question39Pivot",
    )
    xlsx_pivot._sort_key = original
    wb, ws = finish(path, "Question 39 — 2024 Average Wheat Yield by Municipality")
    header_row = find_row(ws, "Municipality")
    for row in range(header_row + 1, header_row + 4):
        highlight_row(ws, row, 1, 2)
    add_answer_box(
        ws,
        "D3:H3",
        "D4:H6",
        "Louise had 79.6 bu/ac, Rhineland had 78.5 bu/ac, and Morris had 78.0 bu/ac—the three highest municipal averages.",
    )
    for col in ("D", "E", "F", "G", "H"):
        ws.column_dimensions[col].width = 16
    for row in range(4, 7):
        ws.row_dimensions[row].height = 25
    save_finished(wb, path)


def build_q40(mb_rows):
    path = ANSWER_DIR / "q040.xlsx"
    helper = [row for row in mb_rows if row["Variety"] == "AAC BRANDON (BW 932)"]
    write_source_workbook(path, mb_rows, MB_HEADERS, helper, 9)
    original = with_sort_order(list(range(2020, 2026)))
    add_pivot(
        path,
        9,
        helper,
        MB_HEADERS,
        rows=["Year"],
        values=[{
            "field": "Acres",
            "agg": "sum",
            "name": "Total Acres",
            "num_format": "#,##0",
        }],
        name="Question40Pivot",
    )
    xlsx_pivot._sort_key = original
    wb, ws = finish(path, "Question 40 — AAC BRANDON Acres by Year")
    highlight_row(ws, find_row(ws, 2020), 1, 2)
    highlight_row(ws, find_row(ws, 2025), 1, 2)
    add_answer_box(
        ws,
        "D3:H3",
        "D4:H6",
        "Acres fell every year, from 1,644,428 in 2020 to 787,111 in 2025. The most acres were planted in 2020.",
    )
    for col in ("D", "E", "F", "G", "H"):
        ws.column_dimensions[col].width = 16
    for row in range(4, 7):
        ws.row_dimensions[row].height = 25
    save_finished(wb, path)


def build_q41(mb_rows):
    path = ANSWER_DIR / "q041.xlsx"
    varieties = ["AAC BRANDON (BW 932)", "AAC STARBUCK <SECAN>"]
    helper = [row for row in mb_rows if row["Variety"] in varieties]
    write_source_workbook(path, mb_rows, MB_HEADERS, helper, 9)
    original = with_sort_order(list(range(2020, 2026)) + varieties)
    add_pivot(
        path,
        9,
        helper,
        MB_HEADERS,
        rows=["Year"],
        cols=["Variety"],
        values=[{
            "field": "Yield_bu_ac",
            "agg": "average",
            "name": "Average Yield (bu/ac)",
            "num_format": "0.0",
        }],
        name="Question41Pivot",
    )
    xlsx_pivot._sort_key = original
    wb, ws = finish(path, "Question 41 — Average Yield of Two Varieties by Year")
    starbuck_col = find_col(ws, "AAC STARBUCK <SECAN>")
    header_row = find_row(ws, "Year")
    grand_row = find_row(ws, "Grand Total")
    for row in range(header_row + 1, grand_row):
        ws.cell(row, starbuck_col).fill = PatternFill("solid", fgColor=YELLOW)
    add_answer_box(
        ws,
        "F3:J3",
        "F4:J6",
        "AAC STARBUCK <SECAN> had the higher average yield in every year from 2020 through 2025.",
    )
    for col in ("F", "G", "H", "I", "J"):
        ws.column_dimensions[col].width = 16
    for row in range(4, 7):
        ws.row_dimensions[row].height = 25
    save_finished(wb, path)


def main():
    ANSWER_DIR.mkdir(parents=True, exist_ok=True)
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    rm_rows = read_rows(RM_SOURCE, {"Year", "RM", "Yield"})
    mb_rows = read_rows(
        MB_SOURCE, {"Year", "Farms", "Acres", "Yield_bu_ac"}
    )
    builders = {
        33: lambda: build_q33(rm_rows),
        34: lambda: build_q34(rm_rows),
        35: lambda: build_q35(rm_rows),
        36: lambda: build_q36(mb_rows),
        37: lambda: build_q37(mb_rows),
        38: lambda: build_q38(mb_rows),
        39: lambda: build_q39(mb_rows),
        40: lambda: build_q40(mb_rows),
        41: lambda: build_q41(mb_rows),
    }
    requested = [int(value) for value in sys.argv[1:]] or list(builders)
    for question in requested:
        builders[question]()


if __name__ == "__main__":
    main()
