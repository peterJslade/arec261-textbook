#!/usr/bin/env python3
"""Build the answer workbooks for the current Module 1 Questions 31 and 32."""

import csv
import os
import shutil
import sys
import zipfile
from pathlib import Path

sys.path.insert(0, os.path.expanduser("~/.claude/skills/excel-pivot-tables/scripts"))

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
import xlsx_pivot


ROOT = Path(__file__).resolve().parents[1]
SOURCE = ROOT / "practice/data/rm_yields_1990_2025.csv"
ANSWER_DIR = ROOT / "practice/answers"
OUTPUT_DIR = ROOT / "outputs/module01-new-pivot-answers"

GREEN = "4A7C59"
PALE_GREEN = "E2F0D9"
YELLOW = "FFF2CC"
TEXT = "24302A"
THIN_GREEN = Side(style="thin", color="A9C4AE")


def read_rows():
    with SOURCE.open(newline="", encoding="utf-8-sig") as handle:
        rows = list(csv.DictReader(handle))
    for row in rows:
        row["Year"] = int(row["Year"])
        row["RM"] = int(row["RM"])
        row["Yield"] = float(row["Yield"])
    return rows


def write_source_workbook(path, rows, helper_rows):
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    headers = ["Year", "RM", "Crop", "Yield", "Unit"]
    ws.append(headers)
    for row in rows:
        ws.append([row[h] for h in headers])
    for col, header in enumerate(headers, 7):
        ws.cell(1, col, header)
    for row_number, row in enumerate(helper_rows, 2):
        for col, header in enumerate(headers, 7):
            ws.cell(row_number, col, row[header])
    style_data_sheet(ws, len(rows) + 1)
    for col in range(7, 12):
        ws.cell(1, col).fill = PatternFill("solid", fgColor=GREEN)
        ws.cell(1, col).font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
        ws.column_dimensions[ws.cell(1, col).column_letter].hidden = True
    wb.save(path)


def style_data_sheet(ws, last_row):
    header = ws[1]
    for cell in header:
        cell.fill = PatternFill("solid", fgColor=GREEN)
        cell.font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:E{last_row}"
    widths = {"A": 11, "B": 10, "C": 19, "D": 12, "E": 11}
    for column, width in widths.items():
        ws.column_dimensions[column].width = width
    ws.sheet_properties.tabColor = GREEN


def format_pivot_sheet(ws, title):
    ws["A1"] = title
    ws["A1"].font = Font(name="Arial", size=14, bold=True, color=GREEN)
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = GREEN
    ws.freeze_panes = "A4"


def highlight_row(ws, row, first_col, last_col):
    for col in range(first_col, last_col + 1):
        ws.cell(row, col).fill = PatternFill("solid", fgColor=YELLOW)


def add_answer_box(ws, label_cell, label_range, answer_range, answer):
    ws.merge_cells(label_range)
    ws[label_cell] = "Answer to part (b)"
    ws[label_cell].fill = PatternFill("solid", fgColor=GREEN)
    ws[label_cell].font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
    ws[label_cell].alignment = Alignment(horizontal="center")
    ws.merge_cells(answer_range)
    cell = ws[answer_range.split(":")[0]]
    cell.value = answer
    cell.fill = PatternFill("solid", fgColor=YELLOW)
    cell.font = Font(name="Arial", size=11, bold=True, color=TEXT)
    cell.alignment = Alignment(vertical="center", wrap_text=True)
    cell.border = Border(
        left=THIN_GREEN, right=THIN_GREEN, top=THIN_GREEN, bottom=THIN_GREEN
    )


def disable_refresh_on_open(path):
    temp = path.with_suffix(".patched.xlsx")
    with zipfile.ZipFile(path, "r") as source, zipfile.ZipFile(
        temp, "w", zipfile.ZIP_DEFLATED
    ) as target:
        for item in source.infolist():
            data = source.read(item.filename)
            if item.filename.startswith("xl/pivotCache/pivotCacheDefinition"):
                data = data.replace(b'refreshOnLoad="1"', b'refreshOnLoad="0"')
            target.writestr(item, data)
    temp.replace(path)


def build_q31(full_rows):
    path = ANSWER_DIR / "q031.xlsx"
    filtered = [
        row
        for row in full_rows
        if row["Year"] == 2023 and row["Crop"] == "Spring Wheat"
    ]
    filtered.sort(key=lambda row: row["Yield"], reverse=True)
    write_source_workbook(path, full_rows, filtered)
    original_sort_key = xlsx_pivot._sort_key
    rank = {row["RM"]: index for index, row in enumerate(filtered)}
    xlsx_pivot._sort_key = lambda value: (rank[value],) if value in rank else original_sort_key(value)
    xlsx_pivot.add_pivot_table(
        path,
        source_sheet="Data",
        source_ref=f"G1:K{len(filtered) + 1}",
        dest_sheet="PivotTable",
        dest_cell="A3",
        rows=["RM"],
        values=[
            {
                "field": "Yield",
                "agg": "average",
                "name": "Average Yield (bu/ac)",
                "num_format": "0.0",
            }
        ],
        name="Question31Pivot",
        output_path=path,
    )
    xlsx_pivot._sort_key = original_sort_key

    wb = load_workbook(path)
    ws = wb["PivotTable"]
    format_pivot_sheet(ws, "Question 31 — 2023 Spring Wheat Yield by RM")

    header_row = next(
        row
        for row in range(1, ws.max_row + 1)
        if ws.cell(row, 1).value in {"Row Labels", "RM"}
    )
    grand_row = next(
        row
        for row in range(header_row + 1, ws.max_row + 1)
        if ws.cell(row, 1).value == "Grand Total"
    )
    results = [
        (ws.cell(row, 1).value, ws.cell(row, 2).value)
        for row in range(header_row + 1, grand_row)
    ]
    results.sort(key=lambda item: item[1], reverse=True)
    for offset, (rm, yield_value) in enumerate(results, header_row + 1):
        ws.cell(offset, 1, rm)
        ws.cell(offset, 2, yield_value)
        ws.cell(offset, 2).number_format = "0.0"
    for row in range(header_row + 1, header_row + 4):
        highlight_row(ws, row, 1, 2)

    answer = "RM 187 — 74.7 bu/ac; RM 494 — 73.0 bu/ac; RM 338 — 72.8 bu/ac."
    add_answer_box(ws, "D3", "D3:H3", "D4:H5", answer)
    for col in ("D", "E", "F", "G", "H"):
        ws.column_dimensions[col].width = 15
    ws.row_dimensions[4].height = 28
    ws.row_dimensions[5].height = 28
    wb.active = wb.sheetnames.index("PivotTable")
    wb.save(path)
    disable_refresh_on_open(path)
    return path


def build_q32(full_rows):
    path = ANSWER_DIR / "q032.xlsx"
    keep_rms = {1, 2, 31, 32}
    keep_crops = {"Spring Wheat", "Peas", "Canola"}
    filtered = [
        row
        for row in full_rows
        if row["Year"] == 2021
        and row["RM"] in keep_rms
        and row["Crop"] in keep_crops
    ]
    write_source_workbook(path, full_rows, filtered)
    original_sort_key = xlsx_pivot._sort_key
    order = {1: 0, 2: 1, 31: 2, 32: 3,
             "Spring Wheat": 0, "Peas": 1, "Canola": 2}
    xlsx_pivot._sort_key = lambda value: (order[value],) if value in order else original_sort_key(value)
    xlsx_pivot.add_pivot_table(
        path,
        source_sheet="Data",
        source_ref=f"G1:K{len(filtered) + 1}",
        dest_sheet="PivotTable",
        dest_cell="A3",
        rows=["RM"],
        cols=["Crop"],
        values=[
            {
                "field": "Yield",
                "agg": "average",
                "name": "Average Yield (bu/ac)",
                "num_format": "0.0",
            }
        ],
        name="Question32Pivot",
        output_path=path,
    )
    xlsx_pivot._sort_key = original_sort_key

    wb = load_workbook(path)
    ws = wb["PivotTable"]
    format_pivot_sheet(ws, "Question 32 — 2021 Yields in RMs 1, 2, 31 and 32")
    for row in range(1, ws.max_row + 1):
        if ws.cell(row, 1).value == "Grand Total":
            for col in range(2, ws.max_column + 1):
                if ws.cell(3, col).value == "Spring Wheat":
                    ws.cell(row, col).fill = PatternFill("solid", fgColor=YELLOW)
            grand_row = row
            break
    else:
        raise RuntimeError("Question 32 PivotTable has no Grand Total row")

    answer_row = grand_row + 3
    answer = "Spring Wheat yielded best across these four RMs, averaging 44.9 bu/ac."
    add_answer_box(
        ws,
        f"A{answer_row}",
        f"A{answer_row}:E{answer_row}",
        f"A{answer_row + 1}:E{answer_row + 2}",
        answer,
    )
    ws.row_dimensions[answer_row + 1].height = 26
    ws.row_dimensions[answer_row + 2].height = 26
    wb.active = wb.sheetnames.index("PivotTable")
    wb.save(path)
    disable_refresh_on_open(path)
    return path


def main():
    ANSWER_DIR.mkdir(parents=True, exist_ok=True)
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    rows = read_rows()
    paths = [build_q31(rows), build_q32(rows)]
    for path in paths:
        shutil.copy2(path, OUTPUT_DIR / path.name)
        print(path)


if __name__ == "__main__":
    main()
