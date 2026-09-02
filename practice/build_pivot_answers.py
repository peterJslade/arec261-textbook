#!/usr/bin/env python3
"""Answer workbooks for the PivotTable questions.

openpyxl cannot create a pivot, so these are built with the excel-pivot-tables
skill, which writes the OOXML cache/table parts directly and also fills in the
result cells. Excel opens them with a live field list.

Each workbook carries the source data on a Data tab and the finished pivot on a
PivotTable tab, so a student can see the input and the output together.

The per-question pivot spec (axes, value field, aggregation, crop/year filters)
lives in practice/pivot_spec.json. Edit that, not this file, to change a pivot.

Run:  python3 practice/build_pivot_answers.py
"""
import csv, json, os, shutil, sys
sys.path.insert(0, os.path.expanduser("~/.claude/skills/excel-pivot-tables/scripts"))
from xlsx_pivot import add_pivot_table
from openpyxl import Workbook, load_workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

DATA, OUT = "practice/data", "practice/answers"
FILES = {"rm_long": "rm_yields_1990_2025.csv",
         "mb":      "mb_wheat_varieties.csv",
         "sc":      "statcan_field_crops.csv"}
NUMCOL = {"rm_long": {"Yield"},
          "mb":      {"Farms", "Acres", "Yield_bu_ac"},
          "sc":      {"Seeded_acres", "Yield_bu_ac"}}
VALFIELD = {"rm_long": "Yield", "mb": "Yield_bu_ac", "sc": "Yield_bu_ac"}
AGGFMT = {"Average": "0.00", "Count": "#,##0", "Sum": "#,##0", "Min": "0.00"}

HEAD = PatternFill("solid", fgColor="4A7C59")
HILITE = PatternFill("solid", fgColor="FFF2CC")
NOTE = PatternFill("solid", fgColor="E2F0D9")
f_h = Font(name="Arial", size=10, bold=True, color="FFFFFF")
f_b = Font(name="Arial", size=10, color="24302A")
f_t = Font(name="Arial", size=12, bold=True, color="4A7C59")
CEN = Alignment(horizontal="center")
WRAP = Alignment(vertical="top", wrap_text=True)
THIN_GREEN = Side(style="thin", color="A9C4AE")

DISPLAY_Q = {
    "71": 31, "76": 32, "74": 33, "101": 34, "80": 35,
    "89": 36, "102": 37, "88": 38, "84": 39, "103": 40,
}

ANSWERS = {
    "71": [
        "Unit, then Crop on Rows; Yield shown as Average. Both unit groups remain visible.",
        "Lentils are about 1,314 lb/ac. They look out of line only if their different unit is ignored.",
        "Among the bu/ac crops, Oats are highest at about 71 and Flax is lowest at about 20.",
    ],
    "76": [
        "Crop on Rows, Year on Columns and Yield shown as Average for 2019-2023 in bu/ac.",
        "Canola is lowest in 2021 (about 21.9 bu/ac); Barley is also lowest in 2021 (about 34.8).",
        "A shared low year suggests that conditions common to many RMs affected both crops.",
    ],
    "74": [
        "Year on Rows and Yield shown as Count for Canola.",
        "A typical recent year has about 290 reporting RMs (2023 has 289).",
        "Count is the number of reported yields; Average would be their mean. A lower count means fewer reporting RMs.",
    ],
    "101": [
        "Year on Rows, with Yield shown once as Minimum and once as Average for Canola.",
        "In 2022 the minimum is 1.9 bu/ac and the average is about 35.4 bu/ac.",
        "The minimum reveals the result for the lowest-yielding reporting RM; the average does not show that extreme.",
    ],
    "80": [
        "Year on Rows and Yield shown as Average for Canola, with a line chart based on the summarized values.",
        "The trend rises overall. The sharpest drop is 2020 to 2021, from about 38.2 to 21.9 bu/ac.",
        "The chart makes the trend and abrupt changes easier to see; the PivotTable makes exact annual averages easier to read.",
    ],
    "89": [
        "Variety on Rows, with Yield shown as Average and Count. All varieties remain visible.",
        "Among varieties with Count at least 30, SY MANNESS is highest (about 72.1) and AAC TISDALE is lowest (about 53.9).",
        "The minimum count removes comparisons supported by only a few reported yields.",
    ],
    "102": [
        "Variety on Rows, Year on Columns and Acres shown as Sum; use the visible Grand Total column for overall totals.",
        "AAC Brandon is first (about 6.74 million acres), AAC Starbuck second (2.50 million) and AAC Wheatland third (1.40 million). AAC Brandon declines every year, from about 1.64 million acres in 2020 to 0.79 million in 2025.",
        "Acres are additive quantities. Yield is a per-acre rate, so adding yields would not give a meaningful total.",
    ],
    "88": [
        "Year on Rows, the two varieties on Columns and Yield shown as Average.",
        "AAC Starbuck has the higher reported average in every year from 2020 through 2025.",
        "The repeated pattern is more informative than one year, but it does not prove causation because growing locations and conditions may differ.",
    ],
    "84": [
        "Year on Rows, with Yield shown as Average and Count for 2020-2025.",
        "The lowest average is 2021 (about 49.6 bu/ac); the highest is 2025 (about 68.0). Averages rise after 2021.",
        "The variety mix changes by year. Count shows how many reported yields support each annual average.",
    ],
    "103": [
        "Average and Count of Yield for AAC Brandon, AAC Starbuck and SY MANNESS, with a horizontal bar chart based on the averages.",
        "SY MANNESS is tallest at about 72.1 bu/ac; AAC Brandon is shortest at about 59.6 bu/ac.",
        "The chart makes the ranking easy to see but hides counts and spread. A zero baseline keeps bar lengths proportional.",
    ],
}

HIGHLIGHT_LABELS = {
    "71": {"Lentils", "Oats", "Flax"},
    "76": {"Canola", "Barley"},
    "74": {2023, "2023"},
    "101": {2022, "2022"},
    "80": {2020, "2020", 2021, "2021"},
    "89": {"SY MANNESS", "AAC TISDALE (PT250)"},
    "102": {"AAC BRANDON (BW 932)", "AAC STARBUCK <SECAN>", "AAC WHEATLAND <SECAN>"},
    "88": {2020, "2020", 2021, "2021", 2022, "2022", 2023, "2023", 2024, "2024", 2025, "2025"},
    "84": {2021, "2021", 2025, "2025"},
    "103": {"SY MANNESS", "AAC BRANDON (BW 932)"},
}


def add_answer_sheet(wb, qno):
    ws = wb.create_sheet("Answer", 0)
    ws.sheet_view.showGridLines = False
    ws.merge_cells("A1:D1")
    ws["A1"] = f"Question {DISPLAY_Q[qno]} — Answer"
    ws["A1"].font = Font(name="Arial", size=15, bold=True, color="FFFFFF")
    ws["A1"].fill = HEAD
    ws["A1"].alignment = Alignment(vertical="center")
    ws.row_dimensions[1].height = 28

    ws["A3"] = "Part"
    ws.merge_cells("B3:D3")
    ws["B3"] = "Answer"
    for cell in ws[3]:
        cell.fill = HEAD
        cell.font = f_h
        cell.alignment = CEN

    for row, (part, answer) in enumerate(zip(("(a)", "(b)", "(c)"), ANSWERS[qno]), 4):
        ws.cell(row, 1, part)
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
        ws.cell(row, 2, answer)
        for col in range(1, 5):
            cell = ws.cell(row, col)
            cell.fill = HILITE
            cell.font = Font(name="Arial", size=11, bold=(col == 1), color="24302A")
            cell.alignment = WRAP
            cell.border = Border(top=THIN_GREEN, bottom=THIN_GREEN)
        ws.row_dimensions[row].height = 46 if len(answer) < 125 else 62

    ws.merge_cells("A9:D10")
    ws["A9"] = "The live PivotTable (and chart, where requested) is on the PivotTable worksheet. The highlighted rows there show the values used in the written answers. The Data worksheet contains the source records."
    ws["A9"].fill = NOTE
    ws["A9"].font = f_b
    ws["A9"].alignment = WRAP
    ws["A9"].border = Border(top=THIN_GREEN, bottom=THIN_GREEN,
                              left=THIN_GREEN, right=THIN_GREEN)
    ws.column_dimensions["A"].width = 10
    for col in ("B", "C", "D"):
        ws.column_dimensions[col].width = 28
    ws.freeze_panes = "A3"
    ws.sheet_properties.tabColor = "4A7C59"
    wb.active = 0


def highlight_pivot_rows(ws, qno):
    labels = HIGHLIGHT_LABELS.get(qno, set())
    for row in range(1, ws.max_row + 1):
        if any(ws.cell(row, col).value in labels
               for col in range(1, ws.max_column + 1)):
            for col in range(1, ws.max_column + 1):
                ws.cell(row, col).fill = HILITE


def write_data(path, ds, years=None, crop=None, unit_bu=False,
               varieties=None, min_variety_count=None,
               sort_varieties_by_average=False, nonblank_field=None):
    rows = list(csv.DictReader(open(f"{DATA}/{FILES[ds]}")))
    # A question about one crop ships that crop only, so the pivot the student
    # builds from this Data tab reproduces the answer key exactly.
    if crop:
        rows = [r for r in rows if r.get("Crop", "").strip().lower() == crop.lower()]
    # several questions say "bu/ac" -- the long file also carries lentils in lb/ac,
    # which would sit at the top of any yield ranking as a pure unit artefact
    if unit_bu:
        rows = [r for r in rows if r.get("Unit", "bu/ac").strip() == "bu/ac"]
    # The long RM file is 71k rows and the pivot cache copies every record, so a
    # question about two named years ships those years only -- both to keep the
    # workbook small and because that is the slice the question is about.
    if years:
        rows = [r for r in rows if r.get("Year") in years]
    if nonblank_field:
        rows = [r for r in rows if str(r.get(nonblank_field, "")).strip()]
    if varieties:
        keep = set(varieties)
        rows = [r for r in rows if r.get("Variety") in keep]
    if min_variety_count:
        counts = {}
        for r in rows:
            if str(r.get("Yield_bu_ac", "")).strip():
                v = r.get("Variety")
                counts[v] = counts.get(v, 0) + 1
        keep = {v for v, count in counts.items() if count >= min_variety_count}
        rows = [r for r in rows if r.get("Variety") in keep]
    if sort_varieties_by_average:
        totals, counts = {}, {}
        for r in rows:
            value = str(r.get("Yield_bu_ac", "")).strip()
            if value:
                v = r.get("Variety")
                totals[v] = totals.get(v, 0.0) + float(value)
                counts[v] = counts.get(v, 0) + 1
        averages = {v: totals[v] / counts[v] for v in totals}
        rows.sort(key=lambda r: -averages.get(r.get("Variety"), float("-inf")))
    cols = list(rows[0].keys())
    num = NUMCOL[ds]
    wb = Workbook(); ws = wb.active; ws.title = "Data"
    for j, h in enumerate(cols, 1):
        c = ws.cell(1, j, h); c.font = f_h; c.fill = HEAD
    for i, r in enumerate(rows, 2):
        for j, h in enumerate(cols, 1):
            v = r[h]
            if h in num and str(v).strip():
                try: v = float(v)
                except ValueError: pass
            c = ws.cell(i, j, v if str(v).strip() else None)
            c.font = f_b
            if h in num: c.alignment = CEN
    for j, h in enumerate(cols, 1):
        ws.column_dimensions[chr(64 + j)].width = max(11, min(20, len(h) + 4))
    ws.freeze_panes = "A2"
    wb.save(path)
    return len(rows), cols


def build(qno, spec):
    ds = spec["ds"]
    path = f"{OUT}/q{int(qno):03d}.xlsx"
    n, cols = write_data(
        path, ds, spec.get("years") or None, spec.get("crop"),
        spec.get("unit_bu", False), spec.get("varieties"),
        spec.get("min_variety_count"),
        spec.get("sort_varieties_by_average", False),
        spec.get("nonblank_field")
    )

    vf = spec.get("valfield") or VALFIELD[ds]
    values = [{"field": vf, "agg": a.lower(),
               "name": f"{a} of {vf}", "num_format": AGGFMT.get(a, "0.00")}
              for a in spec.get("aggs", ["Average"])]

    # the spec omits empty keys so it diffs cleanly; default them here
    rows = [r for r in spec.get("rows", []) if r in cols]
    kcols = [c for c in spec.get("cols", []) if c in cols]
    filters = [f for f in spec.get("filters", []) if f in cols and f not in rows + kcols]
    if not rows and not kcols:
        rows = [c for c in cols if c not in NUMCOL[ds]][:1]

    add_pivot_table(path, source_sheet="Data", dest_sheet="PivotTable",
                    rows=rows, cols=kcols, values=values, filters=filters,
                    output_path=path)

    wb = load_workbook(path)
    ws = wb["PivotTable"]

    chart_type = spec.get("chart")
    if chart_type:
        # A pivot with multiple value fields inserts an extra "Values" row, so
        # locate the real row-label header instead of assuming it is row 3.
        header_row = next(
            row for row in range(2, ws.max_row + 1)
            if ws.cell(row, 1).value in {"Row Labels", rows[0]}
            and ws.cell(row, 2).value is not None
        )
        last_row = ws.max_row
        for row in range(header_row + 1, ws.max_row + 1):
            if ws.cell(row, 1).value == "Grand Total":
                last_row = row - 1
                break

        # Use a small hidden chart source.  It keeps the displayed PivotTable
        # live while allowing the ranked bar chart to be ordered by its value.
        chart_rows = [
            (ws.cell(row, 1).value, ws.cell(row, 2).value)
            for row in range(header_row + 1, last_row + 1)
        ]
        if chart_type == "bar" and spec.get("sort_varieties_by_average"):
            chart_rows.sort(key=lambda pair: pair[1], reverse=True)
        chart_ws = wb.create_sheet("_ChartData")
        chart_ws.sheet_state = "hidden"
        chart_ws.append([ws.cell(header_row, 1).value,
                         ws.cell(header_row, 2).value])
        for pair in chart_rows:
            chart_ws.append(pair)

        chart = LineChart() if chart_type == "line" else BarChart()
        if chart_type == "bar":
            chart.type = "bar"
            chart.style = 10
        data = Reference(chart_ws, min_col=2, min_row=1,
                         max_row=len(chart_rows) + 1)
        cats = Reference(chart_ws, min_col=1, min_row=2,
                         max_row=len(chart_rows) + 1)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.title = spec.get("chart_title", "PivotChart")
        if chart_type == "bar":
            # In openpyxl's horizontal bar chart, x_axis is the category
            # (vertical) axis and y_axis is the value (horizontal) axis.
            chart.x_axis.title = spec.get("y_axis_title", "")
            chart.y_axis.title = spec.get("x_axis_title", "")
        else:
            chart.x_axis.title = spec.get("x_axis_title", "")
            chart.y_axis.title = spec.get("y_axis_title", "")
        chart.legend = None
        chart.height = 9
        chart.width = 16
        ws.add_chart(chart, "E3")
    highlight_pivot_rows(ws, qno)
    add_answer_sheet(wb, qno)
    wb.save(path)
    os.makedirs("outputs/module01-pivot-answers", exist_ok=True)
    shutil.copy2(path, f"outputs/module01-pivot-answers/{os.path.basename(path)}")
    return path, n, rows, kcols, filters, spec.get("aggs", ["Average"])


if __name__ == "__main__":
    spec = json.load(open(os.path.join(os.path.dirname(__file__), "pivot_spec.json")))
    requested = set(sys.argv[1:])
    todo = {k: v for k, v in spec.items()
            if not v.get("skip") and (not requested or k in requested)}
    print("  building %d pivot answer workbooks\n" % len(todo))
    for q in sorted(todo, key=int):
        p, n, r, c, f, a = build(q, todo[q])
        print("   q%-4s %-8s rows=%-14s cols=%-8s filt=%-10s %s (%d src rows)"
              % (q, todo[q]["ds"], ",".join(r)[:14], ",".join(c)[:8],
                 ",".join(f)[:10], "+".join(a), n))
