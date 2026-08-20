#!/usr/bin/env python3
"""Answer workbooks for the Module 1 test bank.

One workbook per question in practice/answers/. Each holds the sheet a correct
answer produces: the relevant slice of data, the formulas live so they recompute,
and a FORMULATEXT column showing what was typed.

Sections 2-4 work on real extracts (a year of RM data, a year of Manitoba
variety data, and so on). The workbook carries that extract on a Data tab and
puts the statistics on an Answer tab that points at it, so a student can see
both the numbers and the formulas that produced them.

Edit this script, not the .xlsx files it writes.

Run:  python3 practice/build_answers.py            # all sections
      python3 practice/build_answers.py 2          # one section
"""

import csv
import os
import sys
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

PRAIRIE = "4A7C59"; INK = "24302A"; MUTED = "5C6B62"; LOCKFIL = "F6EFD9"
FT = "_xlfn.FORMULATEXT"
QI = "_xlfn.QUARTILE.INC"
PI = "_xlfn.PERCENTILE.INC"
SD = "_xlfn.STDEV.S"
VR = "_xlfn.VAR.S"

DATA = "practice/data"
OUT = "practice/answers"

head_fill = PatternFill("solid", fgColor=PRAIRIE)
lock_fill = PatternFill("solid", fgColor=LOCKFIL)
f_title = Font(name="Arial", size=13, bold=True, color=PRAIRIE)
f_sub   = Font(name="Arial", size=10, color=MUTED)
f_head  = Font(name="Arial", size=11, bold=True, color="FFFFFF")
f_body  = Font(name="Arial", size=10, color=INK)
f_bold  = Font(name="Arial", size=10, bold=True, color=INK)
f_form  = Font(name="Consolas", size=9, color="2F5D46")
f_note  = Font(name="Arial", size=9, color=MUTED)
centre  = Alignment(horizontal="center")
wrap    = Alignment(horizontal="left", vertical="top", wrap_text=True)


def put(ws, r, c, v, font=f_body, fill=None, fmt=None, align=None):
    cell = ws.cell(r, c, v)
    cell.font = font
    if fill:  cell.fill = fill
    if fmt:   cell.number_format = fmt
    if align: cell.alignment = align
    return cell


def header(ws, r, labels):
    for i, lab in enumerate(labels, start=1):
        put(ws, r, i, lab, font=f_head, fill=head_fill,
            align=centre if i > 1 else None)


def note(ws, r, text, ncols=6):
    put(ws, r, 1, text, font=f_note)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncols)
    ws.row_dimensions[r].height = 30


def widths(ws, pairs):
    for col, w in pairs:
        ws.column_dimensions[col].width = w


def heights(ws, last):
    for r in range(1, last + 2):
        if ws.row_dimensions[r].height is None:
            ws.row_dimensions[r].height = 18


def book(title, subtitle, ncols=6):
    wb = Workbook(); ws = wb.active; ws.title = "Answer"
    put(ws, 1, 1, title, font=f_title)
    c = put(ws, 2, 1, subtitle, font=f_sub); c.alignment = wrap
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
    ws.row_dimensions[2].height = 32
    return wb, ws


def save(wb, qno):
    os.makedirs(OUT, exist_ok=True)
    p = f"{OUT}/q{qno:03d}.xlsx"
    wb.save(p)
    return p


# ---------------------------------------------------------------------------
# data
# ---------------------------------------------------------------------------
def load(name):
    return list(csv.DictReader(open(f"{DATA}/{name}")))


def data_tab(wb, rows, cols, title):
    """Write an extract onto its own tab and return (sheet name, first, last)."""
    ws = wb.create_sheet("Data")
    put(ws, 1, 1, title, font=f_title)
    header(ws, 3, cols)
    for i, r in enumerate(rows):
        for j, c in enumerate(cols):
            v = r.get(c, "")
            if isinstance(v, str) and v.strip():
                try: v = float(v)
                except ValueError: pass
            else:
                v = None
            put(ws, 4 + i, 1 + j, v,
                fmt="0.0" if isinstance(v, float) and abs(v) < 10000 else None,
                align=centre if j else None)
    last = 3 + len(rows)
    widths(ws, [(get_column_letter(i + 1), 15) for i in range(len(cols))])
    ws.freeze_panes = "A4"
    return "Data", 4, last


def stat_block(ws, r0, sheet, col_letter, first, last, items, label_w=30):
    """A labelled column of statistics pointing at the Data tab."""
    rng = f"'{sheet}'!{col_letter}{first}:{col_letter}{last}"
    r = r0
    for label, formula, fmt in items:
        put(ws, r, 1, label, font=f_bold)
        put(ws, r, 2, formula.replace("@", rng), fill=lock_fill,
            fmt=fmt, align=centre)
        put(ws, r, 3, f"={FT}(B{r})", font=f_form)
        r += 1
    return r - 1


CORE = [
    ("Count (n)",            "=COUNT(@)",            "#,##0"),
    ("Mean",                 "=AVERAGE(@)",          "0.00"),
    ("Median",               "=MEDIAN(@)",           "0.00"),
    ("Standard deviation",   f"={SD}(@)",            "0.00"),
    ("Variance",             f"={VR}(@)",            "0.00"),
    ("Coefficient of var.",  "=" + SD + "(@)/AVERAGE(@)", "0.000"),
    ("Minimum",              "=MIN(@)",              "0.0"),
    ("Maximum",              "=MAX(@)",              "0.0"),
    ("Range",                "=MAX(@)-MIN(@)",       "0.00"),
    ("Q1",                   f"={QI}(@,1)",          "0.00"),
    ("Q3",                   f"={QI}(@,3)",          "0.00"),
    ("IQR",                  f"={QI}(@,3)-{QI}(@,1)", "0.00"),
    ("10th percentile",      f"={PI}(@,0.1)",        "0.00"),
    ("90th percentile",      f"={PI}(@,0.9)",        "0.00"),
]


def descriptive(qno, title, subtitle, rows, cols, stat_col, tail_note):
    """A Section 2 style workbook: extract on Data, statistics on Answer."""
    wb, ws = book(title, subtitle)
    sheet, first, last = data_tab(wb, rows, cols, title)
    j = cols.index(stat_col) + 1
    put(ws, 4, 1, f"Statistics for {stat_col}", font=f_bold)
    end = stat_block(ws, 5, sheet, get_column_letter(j), first, last, CORE)
    note(ws, end + 2, tail_note, 4)
    widths(ws, [("A", 24), ("B", 14), ("C", 34), ("D", 4)])
    heights(ws, end + 2)
    return wb


if __name__ == "__main__":
    which = sys.argv[1] if len(sys.argv) > 1 else "all"
    print("  (builders are registered per section; see build_s1_answers.py for Section 1)")
