#!/usr/bin/env python3
"""Answer workbooks for Test Bank Section 1 — References, Formatting, Operations.

One workbook per question, written to practice/answers/. Each holds the finished
sheet a correct answer produces: the data, the live formulas, and a FORMULATEXT
column showing what was typed.

The data here is synthetic and deliberately tiny -- five or six rows a student can
take in at a glance. This section tests spreadsheet mechanics (anchoring a
reference, formatting a date, getting the order of operations right), and real
survey data would add noise that has nothing to do with those skills. The
statistics sections use the real datasets.

Edit this script, not the .xlsx files it writes. That is the opposite of the rule
for the teaching workbooks in textbook_examples/, which are hand-tuned and are
their own source of truth.

Run:  python3 practice/build_s1_answers.py
"""

import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

PRAIRIE = "4A7C59"; INK = "24302A"; MUTED = "5C6B62"; LOCKFIL = "F6EFD9"
FT = "_xlfn.FORMULATEXT"

head_fill = PatternFill("solid", fgColor=PRAIRIE)
lock_fill = PatternFill("solid", fgColor=LOCKFIL)
f_title = Font(name="Arial", size=13, bold=True, color=PRAIRIE)
f_sub   = Font(name="Arial", size=10, color=MUTED)
f_head  = Font(name="Arial", size=11, bold=True, color="FFFFFF")
f_body  = Font(name="Arial", size=10, color=INK)
f_bold  = Font(name="Arial", size=10, bold=True, color=INK)
f_form  = Font(name="Consolas", size=9, color="2F5D46")
f_note  = Font(name="Arial", size=9, color=MUTED)
centre = Alignment(horizontal="center")
wrap   = Alignment(horizontal="left", vertical="top", wrap_text=True)

OUT = "practice/answers"

# --- the synthetic data, small enough to read at a glance -------------------
FIELDS = [("North", 120, 41.2), ("South", 240, 38.6), ("Creek", 95, 44.8),
          ("Home", 310, 36.1), ("Rented", 175, 39.9)]

DELIVERIES = [("2026-09-14", "Canola", 42.5), ("2026-09-18", "Canola", 38.0),
              ("2026-09-23", "Wheat", 61.2), ("2026-10-02", "Wheat", 55.8),
              ("2026-10-09", "Barley", 47.3)]

CROPS = [("Canola", 14.20), ("Wheat", 8.35), ("Barley", 5.60)]


def put(ws, r, c, v, font=f_body, fill=None, fmt=None, align=None):
    cell = ws.cell(r, c, v)
    cell.font = font
    if fill:  cell.fill = fill
    if fmt:   cell.number_format = fmt
    if align: cell.alignment = align
    return cell


def header(ws, r, labels, first_left=True):
    for i, lab in enumerate(labels, start=1):
        put(ws, r, i, lab, font=f_head, fill=head_fill,
            align=None if (i == 1 and first_left) else centre)


def finish(ws, widths, last_row):
    for col, w in widths:
        ws.column_dimensions[col].width = w
    for r in range(1, last_row + 2):
        if ws.row_dimensions[r].height is None:
            ws.row_dimensions[r].height = 18
    ws.sheet_view.showGridLines = True


def new_book(title, subtitle, ncols=6):
    wb = Workbook(); ws = wb.active; ws.title = "Answer"
    put(ws, 1, 1, title, font=f_title)
    c = put(ws, 2, 1, subtitle, font=f_sub); c.alignment = wrap
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
    ws.row_dimensions[2].height = 32
    return wb, ws


def note(ws, r, text, ncols=6):
    put(ws, r, 1, text, font=f_note)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncols)
    ws.row_dimensions[r].height = 30


def save(wb, qno):
    os.makedirs(OUT, exist_ok=True)
    p = f"{OUT}/s1_q{qno:02d}.xlsx"
    wb.save(p)
    return p


# ---------------------------------------------------------------------------
def q1():
    """One locked price, filled down."""
    wb, ws = new_book(
        "Q1 — Revenue per acre with a locked price",
        "The canola price sits in one cell. Each row multiplies its own yield by "
        "that one cell, so the reference to the price must be absolute. Column E "
        "shows what was typed into column D.")
    put(ws, 4, 1, "Canola price ($/bu)", font=f_bold)
    put(ws, 4, 2, 14.20, font=f_bold, fill=lock_fill, fmt='"$"#,##0.00', align=centre)

    header(ws, 6, ["Field", "Acres", "Yield (bu/ac)", "Revenue ($/ac)", "The formula in D"])
    for i, (name, acres, yld) in enumerate(FIELDS):
        r = 7 + i
        put(ws, r, 1, name)
        put(ws, r, 2, acres, fmt="#,##0", align=centre)
        put(ws, r, 3, yld, fmt="0.0", align=centre)
        put(ws, r, 4, f"=C{r}*$B$4", fmt='"$"#,##0.00', align=centre)
        put(ws, r, 5, f"={FT}(D{r})", font=f_form)
    last = 6 + len(FIELDS)
    note(ws, last + 2,
         "With =C7*B4 instead, filling down walks the price to B5, B6, B7 -- empty "
         "cells -- so every revenue below the first comes out as $0.00.", 5)
    finish(ws, [("A", 12), ("B", 9), ("C", 15), ("D", 15), ("E", 26)], last + 2)
    return wb


# ---------------------------------------------------------------------------
def q2():
    """Mixed references filling a grid."""
    wb, ws = new_book(
        "Q2 — Three crops, one formula",
        "Prices run across row 4; yields run down. A single formula fills the whole "
        "revenue block: the price row is locked with B$4, the yield column is not.", 8)

    put(ws, 4, 1, "Price ($/bu)", font=f_bold)
    for j, (crop, price) in enumerate(CROPS):
        put(ws, 4, 2 + j, price, font=f_bold, fill=lock_fill,
            fmt='"$"#,##0.00', align=centre)
    header(ws, 5, ["Field"] + [c for c, _ in CROPS])
    yields = [(41.2, 52.4, 63.1), (38.6, 49.8, 58.7), (44.8, 55.2, 66.4),
              (36.1, 47.3, 61.9), (39.9, 51.0, 60.2)]
    for i, (name, _, _) in enumerate(FIELDS):
        r = 6 + i
        put(ws, r, 1, name)
        for j in range(3):
            put(ws, r, 2 + j, yields[i][j], fmt="0.0", align=centre)
    last = 5 + len(FIELDS)

    put(ws, 5, 6, "Revenue ($/ac)", font=f_bold)
    for j, (crop, _) in enumerate(CROPS):
        put(ws, 5, 6 + j, crop, font=f_head, fill=head_fill, align=centre)
    for i in range(len(FIELDS)):
        r = 6 + i
        for j in range(3):
            col = chr(ord('B') + j)
            put(ws, r, 6 + j, f"={col}{r}*{col}$4", fmt='"$"#,##0.00', align=centre)
    put(ws, last + 1, 6, f"={FT}(F6)", font=f_form)

    note(ws, last + 3,
         "One formula, =B6*B$4, fills all fifteen cells. Only the price ROW is "
         "locked; the column moves so each crop meets its own price.", 8)
    finish(ws, [("A", 11), ("B", 10), ("C", 9), ("D", 9), ("E", 3),
                ("F", 11), ("G", 11), ("H", 11)], last + 3)
    return wb


# ---------------------------------------------------------------------------
def q3():
    """Typing a table in, with dates and currency formatted."""
    wb, ws = new_book(
        "Q3 — Entering a delivery log",
        "The table was typed in from the question. Dates are real dates (right "
        "aligned, formatted DD-Mon-YYYY), tonnes show one decimal, and the total "
        "row uses SUM.")
    header(ws, 4, ["Date", "Crop", "Tonnes"])
    import datetime as dt
    for i, (d, crop, t) in enumerate(DELIVERIES):
        r = 5 + i
        put(ws, r, 1, dt.date.fromisoformat(d), fmt="DD-MMM-YYYY", align=centre)
        put(ws, r, 2, crop)
        put(ws, r, 3, t, fmt="0.0", align=centre)
    last = 4 + len(DELIVERIES)
    put(ws, last + 1, 1, "Total", font=f_bold)
    put(ws, last + 1, 3, f"=SUM(C5:C{last})", font=f_bold, fill=lock_fill,
        fmt="0.0", align=centre)
    put(ws, last + 1, 4, f"={FT}(C{last+1})", font=f_form)

    note(ws, last + 3,
         "A date typed as text sits on the LEFT of its cell and cannot be used in "
         "arithmetic. If yours is left aligned, Excel did not read it as a date.", 4)
    finish(ws, [("A", 14), ("B", 10), ("C", 10), ("D", 24)], last + 3)
    return wb


# ---------------------------------------------------------------------------
def q4():
    """Date arithmetic."""
    wb, ws = new_book(
        "Q4 — Days between deliveries",
        "Subtracting one date from another gives a number of days, because Excel "
        "stores dates as numbers. The result cell must be formatted as a NUMBER -- "
        "left as a date it shows a day in January 1900.")
    header(ws, 4, ["Date", "Crop", "Days since previous", "The formula in C"])
    import datetime as dt
    for i, (d, crop, t) in enumerate(DELIVERIES):
        r = 5 + i
        put(ws, r, 1, dt.date.fromisoformat(d), fmt="DD-MMM-YYYY", align=centre)
        put(ws, r, 2, crop)
        if i:
            put(ws, r, 3, f"=A{r}-A{r-1}", fmt="0", align=centre)
            put(ws, r, 4, f"={FT}(C{r})", font=f_form)
        else:
            put(ws, r, 3, "—", align=centre)
    last = 4 + len(DELIVERIES)
    put(ws, last + 2, 1, "First to last", font=f_bold)
    put(ws, last + 2, 3, f"=A{last}-A5", font=f_bold, fill=lock_fill, fmt="0", align=centre)
    note(ws, last + 4,
         "25 days from the first delivery to the last. If a gap shows as a date "
         "like 04-Jan-1900, the cell is still formatted as a date -- change it to "
         "Number.", 4)
    finish(ws, [("A", 14), ("B", 10), ("C", 20), ("D", 22)], last + 4)
    return wb


# ---------------------------------------------------------------------------
def q5():
    """Order of operations."""
    wb, ws = new_book(
        "Q5 — Where the brackets go",
        "Four ways of writing the same intended calculation: revenue per acre after "
        "a $40/ac input cost, times the acres. Only one is right.")
    put(ws, 4, 1, "Yield (bu/ac)", font=f_bold); put(ws, 4, 2, 41.2, fill=lock_fill, fmt="0.0", align=centre)
    put(ws, 5, 1, "Price ($/bu)", font=f_bold);  put(ws, 5, 2, 14.20, fill=lock_fill, fmt='"$"#,##0.00', align=centre)
    put(ws, 6, 1, "Cost ($/ac)", font=f_bold);   put(ws, 6, 2, 40.00, fill=lock_fill, fmt='"$"#,##0.00', align=centre)
    put(ws, 7, 1, "Acres", font=f_bold);         put(ws, 7, 2, 120, fill=lock_fill, fmt="#,##0", align=centre)

    header(ws, 9, ["", "Formula", "Result", "Right?"])
    variants = [
        ("Correct", "=(B4*B5-B6)*B7", "Net per acre first, then scaled by acres."),
        ("No brackets", "=B4*B5-B6*B7", "Excel multiplies before subtracting, so this "
                                        "subtracts the whole cost bill from per-acre revenue."),
        ("Brackets in the wrong place", "=B4*(B5-B6)*B7", "Subtracts the cost from the "
                                        "PRICE, not from the revenue."),
        ("Cost applied twice", "=(B4*B5-B6)*B7-B6", "Charges the input cost again on top."),
    ]
    for i, (label, formula, why) in enumerate(variants):
        r = 10 + i
        put(ws, r, 1, label, font=f_bold if i == 0 else f_body)
        put(ws, r, 2, formula, fmt='"$"#,##0.00', align=centre)
        put(ws, r, 3, f"={FT}(B{r})", font=f_form)
        put(ws, r, 5, why, font=f_note)
    last = 9 + len(variants)
    note(ws, last + 2,
         "The correct answer is $65,405 -- (41.2 x $14.20 - $40) x 120. PEMDAS means "
         "multiplication happens before subtraction unless brackets say otherwise.", 5)
    finish(ws, [("A", 26), ("B", 18), ("C", 22), ("D", 3), ("E", 52)], last + 2)
    return wb


if __name__ == "__main__":
    for n, fn in enumerate((q1, q2, q3, q4, q5), start=1):
        print("  wrote", save(fn(), n))
