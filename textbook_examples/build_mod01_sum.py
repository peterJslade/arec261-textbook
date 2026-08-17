#!/usr/bin/env python3
"""Build mod01_sum.xlsx — a small worksheet illustrating ranges and the SUM function.

One tab, kept short so it fits on screen in the 700px embed. Column D shows the
formula that produced column C, using FORMULATEXT.

Run:  python3 textbook_examples/build_mod01_sum.py
"""

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

PRAIRIE = "4A7C59"
INK     = "24302A"
MUTED   = "5C6B62"
FORMULA = "2F5D46"
LOCKFIL = "F6EFD9"

FT = "_xlfn.FORMULATEXT"

head_fill = PatternFill("solid", fgColor=PRAIRIE)
lock_fill = PatternFill("solid", fgColor=LOCKFIL)

f_title = Font(name="Arial", size=14, bold=True, color=PRAIRIE)
f_sub   = Font(name="Arial", size=10, color=MUTED)
f_head  = Font(name="Arial", size=11, bold=True, color="FFFFFF")
f_body  = Font(name="Arial", size=10, color=INK)
f_bold  = Font(name="Arial", size=10, bold=True, color=INK)
f_note  = Font(name="Arial", size=10, color=MUTED)
f_form  = Font(name="Consolas", size=10, color=FORMULA)

centre = Alignment(horizontal="center")
left   = Alignment(horizontal="left")
wrap   = Alignment(horizontal="left", vertical="top", wrap_text=True)

wb = Workbook()
ws = wb.active
ws.title = "SUM example"


def put(r, c, v, font=f_body, fill=None, fmt=None, align=None):
    cell = ws.cell(r, c, v)
    cell.font = font
    if fill:  cell.fill = fill
    if fmt:   cell.number_format = fmt
    if align: cell.alignment = align
    return cell


ws.cell(1, 1, "Ranges and the SUM function").font = f_title
c = ws.cell(2, 1,
    "Five fields of wheat.  Column B is the size of each field in acres and column C "
    "is its yield in bushels per acre.  Multiplying the two gives the bushels each "
    "field produced (column D).  To add those up we hand SUM a range — D5:D9 — rather "
    "than writing D5+D6+D7+D8+D9.")
c.font = f_sub; c.alignment = wrap
ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=5)
ws.row_dimensions[2].height = 44

# --- the data --------------------------------------------------------------
for i, lab in enumerate(["Field", "Acres", "Yield (bu/ac)", "Bushels", "The formula in D"]):
    put(4, i + 1, lab, font=f_head, fill=head_fill,
        align=left if i in (0, 4) else centre)

FIRST = 5
fields = [("North", 120, 52.4), ("South", 240, 48.9), ("Creek", 95, 61.2),
          ("Home",  310, 55.7), ("Rented", 175, 44.3)]
for i, (name, acres, yld) in enumerate(fields):
    r = FIRST + i
    put(r, 1, name)
    put(r, 2, acres, fmt="#,##0", align=centre)
    put(r, 3, yld, fmt="0.0", align=centre)
    put(r, 4, f"=B{r}*C{r}", fmt="#,##0", align=centre)
    put(r, 5, f"={FT}(D{r})", font=f_form)
LAST = FIRST + len(fields) - 1

# --- the totals ------------------------------------------------------------
TOT = LAST + 1
put(TOT, 1, "Total", font=f_bold)
put(TOT, 2, f"=SUM(B{FIRST}:B{LAST})", font=f_bold, fill=lock_fill,
    fmt="#,##0", align=centre)
put(TOT, 4, f"=SUM(D{FIRST}:D{LAST})", font=f_bold, fill=lock_fill,
    fmt="#,##0", align=centre)
put(TOT, 5, f"={FT}(D{TOT})", font=f_form)

# --- a few more ways to write a range --------------------------------------
EX = TOT + 2
put(EX, 1, "Other ways to give SUM a range", font=f_bold)
rows = [
    ("One column",            f"=SUM(B{FIRST}:B{LAST})",
     "The five acre figures, B5 down to B9."),
    ("One row",               f"=SUM(B{FIRST}:D{FIRST})",
     "Straight across row 5 — acres, yield and bushels. Rarely what you want."),
    ("Two ranges at once",    f"=SUM(B{FIRST}:B{LAST},D{FIRST}:D{LAST})",
     "Separate ranges with a comma."),
    ("A range plus a number", f"=SUM(B{FIRST}:B{LAST},40)",
     "Adds 40 acres of summerfallow that is not in the table."),
]
for i, (label, formula, note) in enumerate(rows):
    r = EX + 1 + i
    put(r, 1, label)
    put(r, 2, formula, fmt="#,##0", align=centre)
    put(r, 3, f"={FT}(B{r})", font=f_form)
    put(r, 5, note, font=f_note)
    # the formula text is wide; let it run across C and D
    ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=4)
LASTROW = EX + len(rows)

for col, w in [("A", 26), ("B", 12), ("C", 14), ("D", 12), ("E", 46)]:
    ws.column_dimensions[col].width = w
for r in range(1, LASTROW + 2):
    if ws.row_dimensions[r].height is None:
        ws.row_dimensions[r].height = 18
ws.freeze_panes = "A5"
ws.sheet_view.showGridLines = True

OUT = "textbook_examples/mod01_sum.xlsx"
wb.save(OUT)
print(f"wrote {OUT}")

# --- checks ---------------------------------------------------------------
s = load_workbook(OUT).active
bare = sum(1 for row in s.iter_rows() for c in row
           if isinstance(c.value, str) and "FORMULATEXT" in c.value
           and "_xlfn." not in c.value)
pop = {c.row for row in s.iter_rows() for c in row if c.value is not None}
missing = [r for r in pop if s.row_dimensions[r].height is None]
print(f"  {max(pop)} rows, bare FORMULATEXT={bare}, rows missing height={len(missing)}")
assert bare == 0, "unprefixed FORMULATEXT"
assert not missing, "rows without an explicit height"
assert max(pop) <= 18, f"{max(pop)} rows is too tall for the embed"

# arithmetic the prose will quote
tot_acres = sum(a for _, a, _ in fields)
tot_bu    = sum(a * y for _, a, y in fields)
print(f"  total acres = {tot_acres:,}   total bushels = {tot_bu:,.0f}")
