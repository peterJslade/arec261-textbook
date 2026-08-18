#!/usr/bin/env python3
"""Build mod01_pivot.xlsx — the PivotTable worksheet for Module 1, chapter 2.

Tabs:
  Start here  — what a PivotTable does, and what to try on the Data tab
  Data        — real SCIC variety data, six major crops, 2021-2025 (long format)
  Summary 1   — average yield by crop            (Crop on Rows, Yield in Values)
  Summary 2   — average yield by crop and year   (Year added to Columns)
  Summary 3   — total acres by crop              (the same pivot, Sum of Acres)

The summary tabs hold the numbers a PivotTable produces, so a student can build
their own from the Data tab and check it. They are written as static values
rather than live PivotTables: openpyxl cannot create a real PivotTable object,
and the embed viewer would not show the field list anyway.

Run:  python3 textbook_examples/build_mod01_pivot.py
"""

import csv
from collections import defaultdict
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.worksheet.hyperlink import Hyperlink

PRAIRIE = "4A7C59"
INK     = "24302A"
MUTED   = "5C6B62"
LOCKFIL = "F6EFD9"

head_fill = PatternFill("solid", fgColor=PRAIRIE)
lock_fill = PatternFill("solid", fgColor=LOCKFIL)

f_title = Font(name="Arial", size=14, bold=True, color=PRAIRIE)
f_sub   = Font(name="Arial", size=10, color=MUTED)
f_head  = Font(name="Arial", size=11, bold=True, color="FFFFFF")
f_body  = Font(name="Arial", size=10, color=INK)
f_bold  = Font(name="Arial", size=10, bold=True, color=INK)
f_note  = Font(name="Arial", size=10, color=MUTED)
f_link  = Font(name="Arial", size=11, bold=True, color="0563C1", underline="single")

centre = Alignment(horizontal="center")
left   = Alignment(horizontal="left")
wrap   = Alignment(horizontal="left", vertical="top", wrap_text=True)

# six crops that between them cover most of Saskatchewan's acres
# All six are reported in bushels per acre.  Pulse crops (lentils, chickpeas,
# fababeans) are reported in POUNDS per acre in this dataset, so mixing them in
# would make an average-yield comparison meaningless.
MAIN = ["Wheat - Hard Red Spring", "Wheat - Durum", "Canola/Rapeseed", "Barley",
        "Field Peas", "Oats"]
SHORT = {"Wheat - Hard Red Spring": "Spring Wheat", "Wheat - Durum": "Durum",
         "Canola/Rapeseed": "Canola"}


def put(ws, r, c, v, font=f_body, fill=None, fmt=None, align=None):
    cell = ws.cell(r, c, v)
    cell.font = font
    if fill:  cell.fill = fill
    if fmt:   cell.number_format = fmt
    if align: cell.alignment = align
    return cell


def head(ws, r, labels):
    for i, lab in enumerate(labels, start=1):
        put(ws, r, i, lab, font=f_head, fill=head_fill,
            align=left if i == 1 else centre)


def finish(ws, widths, last_row, freeze):
    for col, w in widths:
        ws.column_dimensions[col].width = w
    for r in range(1, last_row + 2):
        if ws.row_dimensions[r].height is None:
            ws.row_dimensions[r].height = 18
    ws.freeze_panes = freeze
    ws.sheet_view.showGridLines = True


# --- load and slice -------------------------------------------------------
rows = [r for r in csv.DictReader(open("practice/data/sask_variety_yields.csv"))
        if r["Acres"].strip() and r["Yield"].strip() and r["Crop"] in MAIN]
records = [(int(r["Risk_Zone"]), SHORT.get(r["Crop"], r["Crop"]), r["Variety"],
            int(r["Year"]), float(r["Acres"]), float(r["Yield"])) for r in rows]
records.sort(key=lambda t: (t[3], t[1], t[0]))
crops = sorted({c for _, c, _, _, _, _ in records})
years = sorted({y for _, _, _, y, _, _ in records})

wb = Workbook()
wb.remove(wb.active)

# ==========================================================================
# Start here
# ==========================================================================
ws = wb.create_sheet("Start here")
ws.cell(1, 1, "Summarizing a big table").font = f_title
c = ws.cell(2, 1,
    f"The Data tab holds {len(records):,} rows of real Saskatchewan crop data — one row for "
    "every variety, in every risk zone, in every year from 2021 to 2025.  Far too much to "
    "read.  A PivotTable turns it into a summary small enough to think about.  The other "
    "four tabs are along the bottom of the window: Data, then three summaries.")
c.font = f_sub; c.alignment = wrap
ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=6)
ws.row_dimensions[2].height = 44

put(ws, 4, 1, "Data", font=f_link).hyperlink = Hyperlink(ref="A4", location="'Data'!A1")
put(ws, 4, 2, f"The full table: risk zone, crop, variety, year, acres, yield.")
ws.merge_cells(start_row=4, start_column=2, end_row=4, end_column=6)

items = [
    ("Summary 1", "Average yield by crop.  Crop on Rows, Yield in Values, set to Average."),
    ("Summary 2", "Average yield by crop AND year.  The same, with Year added to Columns."),
    ("Summary 3", "Total acres by crop.  Acres in Values, left on Sum."),
]
for i, (tab, desc) in enumerate(items):
    r = 6 + i * 2
    put(ws, r, 1, tab, font=f_link).hyperlink = Hyperlink(
        ref=f"A{r}", location=f"'{tab}'!A1")
    put(ws, r, 2, desc)
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)

put(ws, 13, 1,
    "Each summary tab shows what a PivotTable produces.  Build your own from the Data tab "
    "and check that you get the same numbers.  Note that Excel sums the Values field by "
    "default — for a yield you almost always want Average instead, and forgetting to change "
    "it is the most common PivotTable mistake.", font=f_note)
ws.merge_cells(start_row=13, start_column=1, end_row=13, end_column=6)
ws.row_dimensions[13].height = 46

finish(ws, [("A", 13), ("B", 20), ("C", 16), ("D", 14), ("E", 12), ("F", 12)], 14, "A4")

# ==========================================================================
# Data
# ==========================================================================
ws = wb.create_sheet("Data")
ws.cell(1, 1, "The data").font = f_title
c = ws.cell(2, 1,
    "One row per variety, risk zone and year.  This is long format: crop and year are "
    "values in their own columns, which is what lets a PivotTable group by them.  All six "
    "crops here are measured in bushels per acre, so averaging across them is fair.")
c.font = f_sub; c.alignment = wrap
ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=6)
ws.row_dimensions[2].height = 30

head(ws, 4, ["Risk zone", "Crop", "Variety", "Year", "Acres", "Yield (bu/ac)"])
for i, (zone, crop, variety, year, acres, yld) in enumerate(records):
    r = 5 + i
    put(ws, r, 1, zone, align=centre)
    put(ws, r, 2, crop)
    put(ws, r, 3, variety)
    put(ws, r, 4, year, align=centre)
    put(ws, r, 5, acres, fmt="#,##0", align=centre)
    put(ws, r, 6, yld, fmt="0.0", align=centre)
DATA_LAST = 4 + len(records)
finish(ws, [("A", 11), ("B", 15), ("C", 22), ("D", 8), ("E", 12), ("F", 14)],
       min(DATA_LAST, 60), "A5")

# ==========================================================================
# helpers for the summaries
# ==========================================================================
def wavg(items):
    """Plain average of the yield values — this is what Excel's PivotTable
    computes when you set Values to Average.  (An acreage-WEIGHTED average would
    be the more meaningful number, and it differs noticeably: oats come out at
    101.2 weighted versus 85.0 plain.  But these tabs exist so a student can
    check their own PivotTable against them, so they have to match what Excel
    actually produces.)"""
    return sum(x[1] for x in items) / len(items) if items else None


by_crop = defaultdict(list)
by_crop_year = defaultdict(list)
acres_by_crop = defaultdict(float)
for zone, crop, variety, year, acres, yld in records:
    by_crop[crop].append((acres, yld))
    by_crop_year[(crop, year)].append((acres, yld))
    acres_by_crop[crop] += acres

# --- Summary 1: average yield by crop -------------------------------------
ws = wb.create_sheet("Summary 1")
ws.cell(1, 1, "Average yield by crop").font = f_title
c = ws.cell(2, 1, "Crop on Rows, Yield in Values, changed from Sum to Average.")
c.font = f_sub
ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=3)
head(ws, 4, ["Crop", "Average yield (bu/ac)"])
for i, crop in enumerate(crops):
    put(ws, 5 + i, 1, crop)
    put(ws, 5 + i, 2, round(wavg(by_crop[crop]), 1), fmt="0.0", align=centre)
finish(ws, [("A", 18), ("B", 22)], 5 + len(crops), "A5")

# --- Summary 2: crop x year ------------------------------------------------
ws = wb.create_sheet("Summary 2")
ws.cell(1, 1, "Average yield by crop and year").font = f_title
c = ws.cell(2, 1, "The same table with Year dragged into Columns.")
c.font = f_sub
ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=7)
head(ws, 4, ["Crop"] + [str(y) for y in years])
for i, crop in enumerate(crops):
    put(ws, 5 + i, 1, crop)
    for j, year in enumerate(years):
        vals = by_crop_year.get((crop, year))
        put(ws, 5 + i, 2 + j,
            round(wavg(vals), 1) if vals else None, fmt="0.0", align=centre)
finish(ws, [("A", 18)] + [(chr(66 + j), 10) for j in range(len(years))],
       5 + len(crops), "A5")

# --- Summary 3: total acres ------------------------------------------------
ws = wb.create_sheet("Summary 3")
ws.cell(1, 1, "Total acres by crop").font = f_title
c = ws.cell(2, 1, "Acres in Values, left on Sum. Five years of acres added together.")
c.font = f_sub
ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=3)
head(ws, 4, ["Crop", "Total acres, 2021-2025"])
for i, crop in enumerate(sorted(acres_by_crop, key=acres_by_crop.get, reverse=True)):
    put(ws, 5 + i, 1, crop)
    put(ws, 5 + i, 2, round(acres_by_crop[crop]), fmt="#,##0", align=centre)
finish(ws, [("A", 18), ("B", 24)], 5 + len(crops), "A5")

OUT = "textbook_examples/mod01_pivot.xlsx"
wb.save(OUT)
print(f"wrote {OUT}")
for ws in load_workbook(OUT).worksheets:
    pop = {c.row for row in ws.iter_rows() for c in row if c.value is not None}
    print(f"  {ws.title:12} {max(pop):5} rows")
print(f"\n  {len(records):,} data rows, {len(crops)} crops, years {years[0]}-{years[-1]}")
