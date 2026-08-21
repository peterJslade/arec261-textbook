#!/usr/bin/env python3
"""Build mod03_merge_excel.xlsx — the three-table merge from the Merging Data chapter,
done with XLOOKUP.  Sheet 1 holds 30 yield rows (three RMs) with the merge formulas
and FORMULATEXT beside them; sheets 2 and 3 hold the full lookup and precipitation
tables.

Run:  python3 textbook_examples/build_mod03_merge.py
"""
import csv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

PRAIRIE, INK, MUTED, FORMULA, LOCKFIL = "4A7C59", "24302A", "5C6B62", "2F5D46", "F6EFD9"
FT, XL = "_xlfn.FORMULATEXT", "_xlfn.XLOOKUP"
head_fill = PatternFill("solid", fgColor=PRAIRIE); flag_fill = PatternFill("solid", fgColor=LOCKFIL)
f_title = Font(name="Arial", size=14, bold=True, color=PRAIRIE)
f_sub = Font(name="Arial", size=10, color=MUTED); f_head = Font(name="Arial", size=11, bold=True, color="FFFFFF")
f_body = Font(name="Arial", size=10, color=INK); f_form = Font(name="Consolas", size=10, color=FORMULA)
wrap = Alignment(horizontal="left", vertical="top", wrap_text=True)

def load(p): return list(csv.reader(open(p)))
yields = load("practice/data/rm_yields_2015_2024.csv")
lookup = load("practice/data/rm_lookup.csv")
precip = load("practice/data/station_precip.csv")
def conv(v):
    try: return float(v) if "." in v else int(v)
    except ValueError: return v if v != "" else None

wb = Workbook()
# ---- sheet 1: merge
ws = wb.active; ws.title = "Merge"
ws.cell(1, 1, "Merging three tables with XLOOKUP").font = f_title
c = ws.cell(2, 1, "Thirty rows of RM yields (three RMs, 2015-2024).  Column I looks up each RM's weather "
    "station from the 'RM lookup' sheet: many yield rows, one lookup row per RM.  Column J looks up the "
    "May-August rain for that station AND that year from the 'Precip' sheet, joining the two keys with "
    "& \"|\" & the way the composite key in Module 1 did.  Column K shows what happens when you forget "
    "the year: XLOOKUP returns the first match for the station, 2015's rain, for every year, and says nothing.")
c.font = f_sub; c.alignment = wrap
ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=12); ws.row_dimensions[2].height = 72
hdr = yields[0] + ["weather_station", "precip_may_aug_mm", "WRONG: station only", "formula (column J)"]
for j, h in enumerate(hdr, 1):
    cell = ws.cell(4, j, h); cell.font = f_head; cell.fill = head_fill
rows = [r for r in yields[1:] if r[1] in ("1", "92", "171")]
first = 5; last = first + len(rows) - 1
nL, nP = len(lookup) - 1, len(precip) - 1
for i, r in enumerate(rows):
    R = first + i
    for j, v in enumerate(r, 1): ws.cell(R, j, conv(v)).font = f_body
    ws.cell(R, 9, f"={XL}(B{R},'RM lookup'!$A$2:$A${nL+1},'RM lookup'!$E$2:$E${nL+1})").font = f_body
    ws.cell(R, 10, f'={XL}(I{R}&"|"&A{R},Precip!$A$2:$A${nP+1}&"|"&Precip!$B$2:$B${nP+1},Precip!$C$2:$C${nP+1})').font = f_body
    ws.cell(R, 11, f"={XL}(I{R},Precip!$A$2:$A${nP+1},Precip!$C$2:$C${nP+1})").font = f_body
    ws.cell(R, 11).fill = flag_fill
    ws.cell(R, 12, f"={FT}(J{R})").font = f_form
ws.cell(last + 2, 9, "Formulas in I and K:").font = f_sub
ws.cell(last + 3, 9, f"={FT}(I{first})").font = f_form
ws.cell(last + 4, 9, f"={FT}(K{first})").font = f_form
ws.cell(last + 6, 1, "Checks").font = Font(name="Arial", size=11, bold=True, color=PRAIRIE)
ws.cell(last + 7, 1, "RMs with no station found").font = f_body
ws.cell(last + 7, 4, f'=COUNTIF(I{first}:I{last},"#N/A")').font = f_body
ws.cell(last + 7, 5, f"={FT}(D{last+7})").font = f_form
ws.cell(last + 8, 1, "Station-years with no rain found").font = f_body
ws.cell(last + 8, 4, f'=COUNTIF(J{first}:J{last},"#N/A")').font = f_body
ws.cell(last + 8, 5, f"={FT}(D{last+8})").font = f_form
for col, w in zip("ABCDEFGHIJKL", (7, 6, 13, 8, 8, 8, 8, 8, 16, 17, 19, 46)): ws.column_dimensions[col].width = w
ws.freeze_panes = "A5"
# ---- sheets 2 and 3
for title, data in (("RM lookup", lookup), ("Precip", precip)):
    s = wb.create_sheet(title)
    for j, h in enumerate(data[0], 1):
        cell = s.cell(1, j, h); cell.font = f_head; cell.fill = head_fill
    for i, r in enumerate(data[1:], 2):
        for j, v in enumerate(r, 1): s.cell(i, j, conv(v)).font = f_body
    for j in range(1, len(data[0]) + 1): s.column_dimensions[chr(64 + j)].width = 18
    s.freeze_panes = "A2"
wb.save("textbook_examples/mod03_merge_excel.xlsx"); print("merge rows", len(rows), "lookup", nL, "precip", nP)
