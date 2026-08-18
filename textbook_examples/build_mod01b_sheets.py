#!/usr/bin/env python3
"""Build the worksheets for Module 1, chapter 2 (Going Further with Excel).

One workbook per topic, each kept short enough to fit the 700px embed:

  mod01_conditional.xlsx  — IF, COUNTIF/SUMIF/AVERAGEIF, and the "&" criterion trick
  mod01_lookup.xlsx       — XLOOKUP and VLOOKUP against a small second table
  mod01_wide_long.xlsx    — the same data in wide and long shape, side by side

All three share one small dataset so a student is not re-learning the data each time:
twelve field records, each with a crop, a year and a yield.

Run:  python3 textbook_examples/build_mod01b_sheets.py
"""

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

PRAIRIE = "4A7C59"
INK     = "24302A"
MUTED   = "5C6B62"
FORMULA = "2F5D46"
LOCKFIL = "F6EFD9"

FT = "_xlfn.FORMULATEXT"
XL = "_xlfn.XLOOKUP"   # XLOOKUP is new enough to need the prefix too

head_fill = PatternFill("solid", fgColor=PRAIRIE)
lock_fill = PatternFill("solid", fgColor=LOCKFIL)

f_title = Font(name="Arial", size=14, bold=True, color=PRAIRIE)
f_sub   = Font(name="Arial", size=10, color=MUTED)
f_head  = Font(name="Arial", size=11, bold=True, color="FFFFFF")
f_body  = Font(name="Arial", size=10, color=INK)
f_bold  = Font(name="Arial", size=10, bold=True, color=INK)
f_note  = Font(name="Arial", size=10, color=MUTED)
f_form  = Font(name="Consolas", size=10, color=FORMULA)

centre = Alignment(horizontal="center")
left   = Alignment(horizontal="left")
wrap   = Alignment(horizontal="left", vertical="top", wrap_text=True)

# one shared dataset: field, crop, year, acres, yield
FIELDS = [
    ("North",   "Canola", 2023, 320, 41.2),
    ("South",   "Wheat",  2023, 250, 58.6),
    ("Creek",   "Canola", 2023, 180, 37.9),
    ("Home",    "Barley", 2023, 210, 72.4),
    ("Rented",  "Wheat",  2023, 400, 61.3),
    ("Airport", "Canola", 2023, 150, 44.8),
    ("Hill",    "Barley", 2023, 275, 68.1),
    ("Slough",  "Wheat",  2023, 330, 55.2),
]


def put(ws, r, c, v, font=f_body, fill=None, fmt=None, align=None):
    cell = ws.cell(r, c, v)
    cell.font = font
    if fill:  cell.fill = fill
    if fmt:   cell.number_format = fmt
    if align: cell.alignment = align
    return cell


def head(ws, r, labels):
    for i, lab in enumerate(labels, start=1):
        put(ws, r, i, lab, font=f_head, fill=head_fill,
            align=left if i == 1 else centre)


def intro(ws, title, text, width):
    ws.cell(1, 1, title).font = f_title
    c = ws.cell(2, 1, text); c.font = f_sub; c.alignment = wrap
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=width)
    ws.row_dimensions[2].height = 44


def finish(ws, widths, last_row, freeze="A5"):
    for col, w in widths:
        ws.column_dimensions[col].width = w
    for r in range(1, last_row + 2):
        if ws.row_dimensions[r].height is None:
            ws.row_dimensions[r].height = 18
    ws.freeze_panes = freeze
    ws.sheet_view.showGridLines = True


def data_block(ws, first_row, cols=("Field", "Crop", "Acres", "Yield (bu/ac)")):
    """Write the shared dataset starting at first_row (header) and return the
    first and last data row numbers."""
    head(ws, first_row, list(cols))
    r0 = first_row + 1
    for i, (fld, crop, _yr, acres, yld) in enumerate(FIELDS):
        r = r0 + i
        put(ws, r, 1, fld)
        put(ws, r, 2, crop, align=centre)
        put(ws, r, 3, acres, fmt="#,##0", align=centre)
        put(ws, r, 4, yld, fmt="0.0", align=centre)
    return r0, r0 + len(FIELDS) - 1


# ==========================================================================
# 1 — CONDITIONAL FUNCTIONS
# ==========================================================================
wb = Workbook(); ws = wb.active; ws.title = "Conditional"
intro(ws, "Conditional functions",
      "Eight fields, each with a crop and a yield.  The functions below all answer "
      "questions about a SUBSET of the rows — only the canola fields, only the yields "
      "above a threshold.  Column F shows the formula that produced column E.", 6)

D0, D1 = data_block(ws, 4)

# IF in a column beside the data
put(ws, 4, 5, "Above 50?", font=f_head, fill=head_fill, align=centre)
put(ws, 4, 6, "The formula in E", font=f_head, fill=head_fill, align=left)
for r in range(D0, D1 + 1):
    put(ws, r, 5, f'=IF(D{r}>50,"Yes","No")', align=centre)
    put(ws, r, 6, f"={FT}(E{r})", font=f_form)

Q = D1 + 2
put(ws, Q, 1, "Questions about a subset", font=f_bold)
rows = [
    ("How many canola fields?",        f'=COUNTIF(B{D0}:B{D1},"Canola")'),
    ("Total canola acres",             f'=SUMIF(B{D0}:B{D1},"Canola",C{D0}:C{D1})'),
    ("Average canola yield",           f'=AVERAGEIF(B{D0}:B{D1},"Canola",D{D0}:D{D1})'),
    ("How many yields above 60?",      f'=COUNTIF(D{D0}:D{D1},">60")'),
    ("Average wheat yield over 250 ac", f'=AVERAGEIFS(D{D0}:D{D1},B{D0}:B{D1},"Wheat",C{D0}:C{D1},">250")'),
    ("How many above the average?",    f'=COUNTIF(D{D0}:D{D1},">"&AVERAGE(D{D0}:D{D1}))'),
]
for i, (label, formula) in enumerate(rows):
    r = Q + 1 + i
    put(ws, r, 1, label)
    put(ws, r, 2, formula, fmt="0.0", align=centre)
    put(ws, r, 3, f"={FT}(B{r})", font=f_form)
    ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=6)
LAST = Q + len(rows)

put(ws, LAST + 2, 1,
    'The last one is the awkward case: the threshold is itself a calculation, so you glue '
    '">" onto it with &.  Quoting the whole thing as ">AVERAGE(...)" would search for that '
    'literal text.', font=f_note)
ws.merge_cells(start_row=LAST + 2, start_column=1, end_row=LAST + 2, end_column=6)
ws.row_dimensions[LAST + 2].height = 32

finish(ws, [("A", 30), ("B", 15), ("C", 11), ("D", 13), ("E", 11), ("F", 26)], LAST + 2)
wb.save("textbook_examples/mod01_conditional.xlsx")
print("wrote mod01_conditional.xlsx")

# ==========================================================================
# 2 — LOOKUPS
# ==========================================================================
wb = Workbook(); ws = wb.active; ws.title = "Lookup"
intro(ws, "Looking values up in another table",
      "The prices live in a separate little table (H4:I7), not in the main one.  XLOOKUP "
      "finds each field's crop in the price list and brings the price back, so we can work "
      "out what the crop is worth.", 6)

D0, D1 = data_block(ws, 4)

# the price table, off to the right
put(ws, 4, 8, "Crop", font=f_head, fill=head_fill, align=centre)
put(ws, 4, 9, "Price $/bu", font=f_head, fill=head_fill, align=centre)
prices = [("Canola", 14.20), ("Wheat", 8.35), ("Barley", 5.60)]
for i, (crop, price) in enumerate(prices):
    r = 5 + i
    put(ws, r, 8, crop, fill=lock_fill, align=centre)
    put(ws, r, 9, price, fill=lock_fill, fmt='"$"#,##0.00', align=centre)

put(ws, 4, 5, "Price $/bu", font=f_head, fill=head_fill, align=centre)
put(ws, 4, 6, "The formula in E", font=f_head, fill=head_fill, align=left)
for r in range(D0, D1 + 1):
    put(ws, r, 5, f"={XL}(B{r},$H$5:$H$7,$I$5:$I$7)", fmt='"$"#,##0.00', align=centre)
    put(ws, r, 6, f"={FT}(E{r})", font=f_form)

Q = D1 + 2
put(ws, Q, 1, "The same lookup, three ways", font=f_bold)
ways = [
    ("XLOOKUP (use this one)", f"={XL}(B{D0},$H$5:$H$7,$I$5:$I$7)"),
    ("VLOOKUP (older files)",  f"=VLOOKUP(B{D0},$H$5:$I$7,2,FALSE)"),
    ("INDEX/MATCH (older still)",
     f"=INDEX($I$5:$I$7,MATCH(B{D0},$H$5:$H$7,0))"),
    ("XLOOKUP, crop not in the list",
     f'={XL}("Lentils",$H$5:$H$7,$I$5:$I$7,"no price")'),
]
for i, (label, formula) in enumerate(ways):
    r = Q + 1 + i
    put(ws, r, 1, label)
    put(ws, r, 2, formula, align=centre)
    put(ws, r, 3, f"={FT}(B{r})", font=f_form)
    ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=6)
LAST = Q + len(ways)

put(ws, LAST + 2, 1,
    "All three return the same number.  XLOOKUP is the one to reach for: it takes the "
    "column to search and the column to return, it matches exactly by default, and the "
    "fourth argument lets you say what to show when nothing is found.", font=f_note)
ws.merge_cells(start_row=LAST + 2, start_column=1, end_row=LAST + 2, end_column=6)
ws.row_dimensions[LAST + 2].height = 32

finish(ws, [("A", 26), ("B", 17), ("C", 13), ("D", 13), ("E", 12), ("F", 24),
            ("G", 3), ("H", 12), ("I", 12)], LAST + 2)
wb.save("textbook_examples/mod01_lookup.xlsx")
print("wrote mod01_lookup.xlsx")

# ==========================================================================
# 3 — WIDE VS LONG  (three tabs: an explanation, then each shape on its own sheet)
# ==========================================================================
wb = Workbook()
wb.remove(wb.active)

wide_rows = [(2023, 1, 50.8, 36.8, 53.0),
             (2023, 2, 48.5, 34.4, 50.5),
             (2023, 3, 52.1, 38.9, 55.7)]

# --- tab 1: the explanation, with links to the other two -------------------
ws = wb.create_sheet("Start here")
ws.cell(1, 1, "The same data in two shapes").font = f_title
c = ws.cell(2, 1,
    "The next two tabs hold exactly the same numbers — three rural municipalities, three "
    "crops, one year — laid out two different ways.  Open them side by side and compare.")
c.font = f_sub; c.alignment = wrap
ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=5)
ws.row_dimensions[2].height = 32

link = Font(name="Arial", size=11, bold=True, color="0563C1", underline="single")

put(ws, 4, 1, "Wide", font=link).hyperlink = "#'Wide'!A1"
put(ws, 4, 2, "One row per RM.  Each crop is its own column, so the crop name is a "
              "column HEADING.", font=f_body)
ws.merge_cells(start_row=4, start_column=2, end_row=4, end_column=5)

put(ws, 6, 1, "Long", font=link).hyperlink = "#'Long'!A1"
put(ws, 6, 2, "One row per RM and crop.  The crop is a VALUE in its own column, and every "
              "yield is stacked into one column.", font=f_body)
ws.merge_cells(start_row=6, start_column=2, end_row=6, end_column=5)

put(ws, 8, 1,
    "Neither is wrong; they suit different jobs.  Wide is easy to read and quick for "
    "column-at-a-time arithmetic — average the Canola column and you are done.  Long is "
    "what you need to summarize BY the stacked thing: a PivotTable can only group by a "
    "field that lives in its own column, so to get average yield per crop, the crop has to "
    "be a column of values rather than a set of headings.", font=f_note)
ws.merge_cells(start_row=8, start_column=1, end_row=8, end_column=5)
ws.row_dimensions[8].height = 60

put(ws, 10, 1,
    "Try it: on the Long tab you can build that PivotTable in one move.  On the Wide tab "
    "there is no Crop field to drag anywhere.", font=f_note)
ws.merge_cells(start_row=10, start_column=1, end_row=10, end_column=5)
ws.row_dimensions[10].height = 32

finish(ws, [("A", 12), ("B", 22), ("C", 18), ("D", 16), ("E", 16)], 11, freeze="A4")

# --- tab 2: wide ------------------------------------------------------------
ws = wb.create_sheet("Wide")
intro(ws, "Wide — one row per RM",
      "Three crops, three columns.  The crop name is a column heading, so it is part of "
      "the layout rather than part of the data.", 5)
head(ws, 4, ["Year", "RM", "Wheat", "Canola", "Barley"])
for i, row in enumerate(wide_rows):
    for j, v in enumerate(row):
        put(ws, 5 + i, 1 + j, v, fmt="0.0" if j >= 2 else "0", align=centre)
finish(ws, [("A", 10), ("B", 8), ("C", 12), ("D", 12), ("E", 12)], 5 + len(wide_rows))

# --- tab 3: long ------------------------------------------------------------
ws = wb.create_sheet("Long")
intro(ws, "Long — one row per RM and crop",
      "The same nine numbers.  Crop is now a column of values and every yield sits in a "
      "single Yield column, which is the shape a PivotTable needs.", 4)
head(ws, 4, ["Year", "RM", "Crop", "Yield"])
r = 5
for (yr, rm, wheat, canola, barley) in wide_rows:
    for crop, val in (("Wheat", wheat), ("Canola", canola), ("Barley", barley)):
        put(ws, r, 1, yr, align=centre)
        put(ws, r, 2, rm, align=centre)
        put(ws, r, 3, crop, align=centre)
        put(ws, r, 4, val, fmt="0.0", align=centre)
        r += 1
finish(ws, [("A", 10), ("B", 8), ("C", 12), ("D", 12)], r - 1)

wb.save("textbook_examples/mod01_wide_long.xlsx")
print("wrote mod01_wide_long.xlsx")

# --- checks ---------------------------------------------------------------
print()
for f in ("mod01_conditional", "mod01_lookup", "mod01_wide_long"):
    s = load_workbook(f"textbook_examples/{f}.xlsx").active
    bare = sum(1 for row in s.iter_rows() for c in row
               if isinstance(c.value, str) and "FORMULATEXT" in c.value
               and "_xlfn." not in c.value)
    pop = {c.row for row in s.iter_rows() for c in row if c.value is not None}
    missing = [r for r in pop if s.row_dimensions[r].height is None]
    print(f"  {f:20} {max(pop):2} rows  bare-FT={bare}  no-height={len(missing)}")
    assert bare == 0, f"{f}: unprefixed FORMULATEXT"
    assert not missing, f"{f}: rows without an explicit height"
    assert max(pop) <= 28, f"{f}: {max(pop)} rows may not fit the embed"
