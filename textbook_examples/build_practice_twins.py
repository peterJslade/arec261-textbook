#!/usr/bin/env python3
"""Make a blank 'practice' twin of each worked teaching workbook.

For every workbook listed below, this writes <name>_practice.xlsx with:
  - every formula cell emptied (the student types them in)
  - the FORMULATEXT columns removed entirely, header included
  - everything else — data, labels, headings, formatting — left alone

Because the twin is derived from the finished workbook rather than written
separately, the two can never drift apart.

Run:  python3 textbook_examples/build_practice_twins.py
"""

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

SOURCES = [
    "mod01_cell_formatting.xlsx",
    "mod01_conditional.xlsx",
    "mod01_lookup.xlsx",
    "mod01_references.xlsx",
    "mod01_sum.xlsx",
]

DIR = "textbook_examples"

# Formula cells to LEAVE IN PLACE. Some formulas are not the exercise: in the
# cell-formatting sheet, B6:B8 are all "=B5" and exist only so the same value can
# be shown under four different number formats. Blanking them would leave nothing
# to format and destroy the point of the sheet.
KEEP = {
    "mod01_cell_formatting.xlsx": {("Formatting", "B6"), ("Formatting", "B7"),
                                   ("Formatting", "B8")},
}


def is_formula(v):
    return isinstance(v, str) and v.startswith("=")


def formulatext_columns(ws):
    """Columns to delete outright: those whose populated cells are ONLY
    FORMULATEXT calls (plus an optional text header). A column that also holds
    data must be kept — otherwise the exercise loses the numbers it needs."""
    ft, other = {}, {}
    for row in ws.iter_rows():
        for c in row:
            if c.value is None:
                continue
            if isinstance(c.value, str) and "FORMULATEXT" in c.value:
                ft[c.column] = ft.get(c.column, 0) + 1
            elif not (isinstance(c.value, str) and c.value.startswith("=")):
                other[c.column] = other.get(c.column, 0) + 1   # data or a label
    keep = []
    for col, n in ft.items():
        # allow at most one non-formula cell (the column heading)
        if other.get(col, 0) <= 1:
            keep.append(col)
    return sorted(keep)


def blank(src_name):
    src = f"{DIR}/{src_name}"
    wb = load_workbook(src)          # formulas, not values
    stats = {"cleared": 0, "cols_removed": 0, "kept": 0}
    keep = KEEP.get(src_name, set())

    for ws in wb.worksheets:
        ft_cols = formulatext_columns(ws)

        # 1. empty every formula cell
        for row in ws.iter_rows():
            for c in row:
                if is_formula(c.value):
                    if (ws.title, c.coordinate) in keep:
                        stats["kept"] += 1
                        continue
                    c.value = None
                    stats["cleared"] += 1

        # 2. drop the FORMULATEXT columns, right to left so indices hold,
        #    taking their header cell with them
        for col in reversed(ft_cols):
            ws.delete_cols(col)
            stats["cols_removed"] += 1
            letter = get_column_letter(col)
            if letter in ws.column_dimensions:
                del ws.column_dimensions[letter]

        # 3. tidy any header left stranded above a now-empty column
        for row in ws.iter_rows(min_row=1, max_row=6):
            for c in row:
                if isinstance(c.value, str) and c.value.strip().lower().startswith(
                        ("the formula", "formula used")):
                    c.value = None

    out = f"{DIR}/{src_name.replace('.xlsx', '_practice.xlsx')}"
    wb.save(out)
    return out, stats


if __name__ == "__main__":
    for name in SOURCES:
        out, st = blank(name)
        print(f"  {out.split('/')[-1]:44} "
              f"cleared {st['cleared']:3}, kept {st['kept']}, "
              f"removed {st['cols_removed']} column(s)")
