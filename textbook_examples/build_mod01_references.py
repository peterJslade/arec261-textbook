#!/usr/bin/env python3
"""Build mod01_references.xlsx — one workbook, three tabs, on cell references.

Each TAB makes one point and fits on screen in the 700px embed (no scrolling):

  Relative       — a relative reference: everything moves together
  Absolute       — one crop, one factor cell, locked with $B$4
  Several crops  — three crops, factors in row 4, locked with C$4

Every sheet pairs a live formula with a FORMULATEXT column so the reader can see
which part of a reference moved and which part stayed put.

Run:  python3 textbook_examples/build_mod01_references.py
"""

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

# --- palette (matches the other Module 1 teaching workbooks) ----------------
PRAIRIE = "4A7C59"   # header fill
PALE    = "EEF4EF"   # banner fill
INK     = "24302A"   # body text
MUTED   = "5C6B62"   # notes
FORMULA = "2F5D46"   # the FORMULATEXT column
LOCKFIL = "F6EFD9"   # pale wheat, for the locked factor cells

# FORMULATEXT is an Excel-2013 function: openpyxl must write the _xlfn. prefix
# or Excel renders #NAME?.
FT = "_xlfn.FORMULATEXT"

# bushel weights (lb/bu) / 2.20462 / 1000  ->  tonnes per bushel
WHEAT_F, BARLEY_F, OATS_F = 0.027216, 0.021772, 0.017237

head_fill = PatternFill("solid", fgColor=PRAIRIE)
lock_fill = PatternFill("solid", fgColor=LOCKFIL)

f_title = Font(name="Arial", size=14, bold=True, color=PRAIRIE)
f_sub   = Font(name="Arial", size=10, color=MUTED)
f_head  = Font(name="Arial", size=11, bold=True, color="FFFFFF")
f_body  = Font(name="Arial", size=10, color=INK)
f_bold  = Font(name="Arial", size=10, bold=True, color=INK)
f_note  = Font(name="Arial", size=10, color=MUTED)
f_form  = Font(name="Consolas", size=10, color=FORMULA)

centre = Alignment(horizontal="center")
left   = Alignment(horizontal="left")
wrap   = Alignment(horizontal="left", vertical="top", wrap_text=True)

LOADS = [500, 1000, 2500, 5000, 8000]


WB = Workbook()
WB.remove(WB.active)          # start clean; every sheet is added explicitly


def new_sheet(title, subtitle, name, width=6):
    """Add a tab to the single shared workbook."""
    ws = WB.create_sheet(name)
    ws.cell(1, 1, title).font = f_title
    c = ws.cell(2, 1, subtitle); c.font = f_sub; c.alignment = wrap
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=width)
    ws.row_dimensions[2].height = 30
    return WB, ws


def put(ws, r, c, v, font=f_body, fill=None, fmt=None, align=None):
    cell = ws.cell(r, c, v)
    cell.font = font
    if fill:  cell.fill = fill
    if fmt:   cell.number_format = fmt
    if align: cell.alignment = align
    return cell


def header(ws, r, labels):
    for i, lab in enumerate(labels, start=1):
        put(ws, r, i, lab, font=f_head, fill=head_fill,
            align=left if i == 1 else centre)


def finish(ws, widths, last_row, freeze):
    for col, w in widths:
        ws.column_dimensions[col].width = w
    for r in range(1, last_row + 2):
        if ws.row_dimensions[r].height is None:
            ws.row_dimensions[r].height = 18
    ws.freeze_panes = freeze
    ws.sheet_view.showGridLines = True   # headers/gridlines ON: the reader needs
                                         # to see which cell is which


# ==========================================================================
# 1 — RELATIVE: every part of the reference moves
# ==========================================================================
wb, ws = new_sheet(
    "Relative references",
    "In this worksheet we have five loads of grain.  Each load has a different size "
    "in bushels (col B) and price per bushel (col C).  We want to calculate the value "
    "of each load by multiplying bushels times price.  To do so, we can just write "
    "B5*C5 in cell D5 and the copy that equation into cells D6-D9.",
    "Ex1 (relative)", width=5)

header(ws, 4, ["Load", "Bushels", "Price ($/bu)", "Value ($)", "The formula in D"])
FIRST = 5
prices = [7.50, 7.50, 8.10, 8.10, 7.95]
for i, bu in enumerate(LOADS):
    r = FIRST + i
    put(ws, r, 1, f"Load {chr(65+i)}")
    put(ws, r, 2, bu, fmt="#,##0", align=centre)
    put(ws, r, 3, prices[i], fmt='"$"#,##0.00', align=centre)
    put(ws, r, 4, f"=B{r}*C{r}", fmt='"$"#,##0', align=centre)
    put(ws, r, 5, f"={FT}(D{r})", font=f_form)
LAST = FIRST + len(LOADS) - 1


finish(ws, [("A", 12), ("B", 12), ("C", 14), ("D", 12), ("E", 26)], LAST + 2, "A5")

# ==========================================================================
# 2 — ABSOLUTE, one crop: one factor cell, locked with $B$4
# ==========================================================================
wb, ws = new_sheet(
    "Absolute references — one crop",
    "Now we want the same five loads in tonnes rather than bushels.  The conversion "
    "factor sits in one cell (B4), and every load needs that same cell.  If we write "
    "B7*B4 and copy it down, the B4 part drifts to B5, B6 and the answers go wrong.  "
    "Writing B7*$B$4 instead locks the factor in place while the bushels still move.",
    "Ex2 (absolute)", width=4)

FACTOR_ROW = 4
put(ws, FACTOR_ROW, 1, "Tonnes per bushel (wheat)", font=f_bold)
put(ws, FACTOR_ROW, 2, WHEAT_F, font=f_bold, fill=lock_fill, fmt="0.00000", align=centre)
put(ws, FACTOR_ROW, 3, "← every formula points here", font=f_note)

header(ws, 6, ["Load", "Bushels", "Tonnes", "The formula in C"])
FIRST = 7
for i, bu in enumerate(LOADS):
    r = FIRST + i
    put(ws, r, 1, f"Load {chr(65+i)}")
    put(ws, r, 2, bu, fmt="#,##0", align=centre)
    put(ws, r, 3, f"=B{r}*$B${FACTOR_ROW}", fmt="#,##0.0", align=centre)
    put(ws, r, 4, f"={FT}(C{r})", font=f_form)
LAST = FIRST + len(LOADS) - 1


finish(ws, [("A", 24), ("B", 12), ("C", 12), ("D", 24)], LAST + 2, "A7")

# ==========================================================================
# 3 — ABSOLUTE, two crops: each factor sits UNDER its own column, locked with B$14
# ==========================================================================
wb, ws = new_sheet(
    "Absolute references — several crops",
    "Now we have two crops, and each has its own conversion factor.  The factors sit in "
    "row 14, directly underneath the column they belong to.  Writing B7*B$14 locks the "
    "row (so every load reads row 14) but leaves the column free — so the wheat column "
    "reads B14 and the barley column reads C14.",
    "Ex3 (absolute)", width=6)

FACTOR_ROW = 14
header(ws, 6, ["Load", "Wheat_bu", "Barley_bu", "Wheat_tons", "Barley_tons",
               "The formula in D"])
FIRST = 7
wheat_bu  = [500, 1000, 2500, 5000, 8000]
barley_bu = [200,  300,  600,  800, 1500]
for i in range(len(wheat_bu)):
    r = FIRST + i
    put(ws, r, 1, f"Load {chr(65+i)}")
    put(ws, r, 2, wheat_bu[i],  fmt="#,##0", align=centre)
    put(ws, r, 3, barley_bu[i], fmt="#,##0", align=centre)
    put(ws, r, 4, f"=B{r}*B${FACTOR_ROW}", fmt="#,##0.0", align=centre)
    put(ws, r, 5, f"=C{r}*C${FACTOR_ROW}", fmt="#,##0.0", align=centre)
    put(ws, r, 6, f"={FT}(D{r})", font=f_form)
LAST = FIRST + len(wheat_bu) - 1

# the factors, each beneath the column it converts
put(ws, FACTOR_ROW, 1, "Tons/bu", font=f_bold)
put(ws, FACTOR_ROW, 2, WHEAT_F,  font=f_bold, fill=lock_fill, fmt="0.00000", align=centre)
put(ws, FACTOR_ROW, 3, BARLEY_F, font=f_bold, fill=lock_fill, fmt="0.00000", align=centre)

finish(ws, [("A", 12), ("B", 12), ("C", 12), ("D", 13), ("E", 13), ("F", 24)],
       FACTOR_ROW, "A7")

# --- save the single workbook ---------------------------------------------
OUT = "textbook_examples/mod01_references.xlsx"
WB.save(OUT)
print(f"wrote {OUT} with sheets: {WB.sheetnames}")

# --- sanity checks ---------------------------------------------------------
print()
for ws in load_workbook(OUT).worksheets:
    bare = sum(1 for row in ws.iter_rows() for c in row
               if isinstance(c.value, str) and "FORMULATEXT" in c.value
               and "_xlfn." not in c.value)
    pop = {c.row for row in ws.iter_rows() for c in row if c.value is not None}
    missing = [r for r in pop if ws.row_dimensions[r].height is None]
    print(f"  {ws.title:16} {max(pop):2} rows  bare-FT={bare}  no-height={len(missing)}")
    assert bare == 0,      f"{ws.title}: unprefixed FORMULATEXT"
    assert not missing,    f"{ws.title}: rows without an explicit height"
    assert max(pop) <= 16, f"{ws.title}: {max(pop)} rows is too tall for the embed"
print("\nthree tabs, each fitting on screen in the 700px embed")
