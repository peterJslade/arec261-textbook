#!/usr/bin/env python3
"""Build mod03_cleaning_excel.xlsx — the grain_deliveries_messy data with Excel
cleaning formulas alongside, each shown with FORMULATEXT.

Run:  python3 textbook_examples/build_mod03_cleaning.py
"""
import csv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

PRAIRIE, INK, MUTED, FORMULA, LOCKFIL = "4A7C59", "24302A", "5C6B62", "2F5D46", "F6EFD9"
FT = "_xlfn.FORMULATEXT"
head_fill = PatternFill("solid", fgColor=PRAIRIE)
flag_fill = PatternFill("solid", fgColor=LOCKFIL)
f_title = Font(name="Arial", size=14, bold=True, color=PRAIRIE)
f_sub   = Font(name="Arial", size=10, color=MUTED)
f_head  = Font(name="Arial", size=11, bold=True, color="FFFFFF")
f_body  = Font(name="Arial", size=10, color=INK)
f_form  = Font(name="Consolas", size=10, color=FORMULA)
wrap = Alignment(horizontal="left", vertical="top", wrap_text=True)

rows = list(csv.DictReader(open("practice/data/grain_deliveries_messy.csv")))
def num(v):
    try: return float(v)
    except ValueError: return None
def defective(r):
    w, m = num(r["weight_tonnes"]), num(r["moisture_pct"])
    return (r["crop"] != r["crop"].strip().title() or w is None or w <= 0
            or m is None or m < 1 or "/" in r["delivery_date"])
keep = [r for r in rows if defective(r)][:14]
ids = {r["ticket_id"] for r in keep}
keep += [r for r in rows if r["ticket_id"] not in ids][:4]
# add one exact duplicate so COUNTIF has something to find
keep.append(dict(keep[2]))

wb = Workbook(); ws = wb.active; ws.title = "Cleaning"
ws.cell(1, 1, "Cleaning a messy file in Excel").font = f_title
c = ws.cell(2, 1, "Grain delivery tickets, as they arrived.  Columns A-F are the raw data.  "
    "Columns G-J are cleaning formulas: a tidied crop name, moisture with the decimal-entered "
    "values rescaled, a duplicate count on the ticket number, and a flag for impossible weights.  "
    "Column K shows the formula in the column to its left.  Nothing in A-F is changed; the "
    "cleaned version sits beside the original so you can still see what you started with.")
c.font = f_sub; c.alignment = wrap
ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=11)
ws.row_dimensions[2].height = 58

hdr = ["ticket_id", "delivery_date", "crop", "weight_tonnes", "moisture_pct", "grade",
       "crop_clean", "moisture_fixed", "ticket_count", "weight_flag", "formula (column J)"]
for j, h in enumerate(hdr, 1):
    cell = ws.cell(4, j, h); cell.font = f_head; cell.fill = head_fill
first, last = 5, 4 + len(keep)
for i, r in enumerate(keep):
    R = first + i
    ws.cell(R, 1, r["ticket_id"]).font = f_body
    ws.cell(R, 2, r["delivery_date"]).font = f_body
    ws.cell(R, 3, r["crop"]).font = f_body
    for col, key in ((4, "weight_tonnes"), (5, "moisture_pct"), (6, "grade")):
        v = r[key]
        try: v = float(v)
        except ValueError: pass
        ws.cell(R, col, v).font = f_body
    ws.cell(R, 7, f"=PROPER(TRIM(C{R}))").font = f_body
    ws.cell(R, 8, f'=IF(ISNUMBER(E{R}),IF(E{R}<1,E{R}*100,E{R}),"")').font = f_body
    ws.cell(R, 9, f"=COUNTIF($A${first}:$A${last},A{R})").font = f_body
    ws.cell(R, 10, f'=IF(D{R}<=0,"check","")').font = f_body
    ws.cell(R, 11, f"={FT}(J{R})").font = f_form
    for col in (9, 10):
        ws.cell(R, col).fill = flag_fill
# formula examples for the other three cleaning columns, below the table
R = last + 2
ws.cell(R, 7, f"={FT}(G{first})").font = f_form
ws.cell(R, 8, f"={FT}(H{first})").font = f_form
ws.cell(R, 9, f"={FT}(I{first})").font = f_form
ws.cell(R - 1, 7, "Formulas in G, H and I:").font = f_sub
R += 2
ws.cell(R, 1, "Summary checks").font = Font(name="Arial", size=11, bold=True, color=PRAIRIE)
checks = [("Smallest weight", f"=MIN(D{first}:D{last})"),
          ("Largest moisture", f"=MAX(E{first}:E{last})"),
          ("Smallest moisture", f"=MIN(E{first}:E{last})"),
          ("Blank moisture readings", f'=COUNTBLANK(E{first}:E{last})'),
          ("Text in the moisture column", f'=SUMPRODUCT(--ISTEXT(E{first}:E{last}))'),
          ("Tickets entered more than once", f"=SUMPRODUCT(--(I{first}:I{last}>1))/2")]
for k, (label, fm) in enumerate(checks):
    ws.cell(R + 1 + k, 1, label).font = f_body
    ws.cell(R + 1 + k, 4, fm).font = f_body
    ws.cell(R + 1 + k, 5, f"={FT}(D{R + 1 + k})").font = f_form
for col, w in zip("ABCDEFGHIJK", (10, 13, 14, 14, 13, 7, 14, 15, 13, 12, 28)):
    ws.column_dimensions[col].width = w
ws.freeze_panes = "A5"
wb.save("textbook_examples/mod03_cleaning_excel.xlsx")
print("rows", len(keep))
